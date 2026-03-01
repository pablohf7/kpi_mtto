import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
from datetime import datetime, timedelta
import re
import hashlib
import time

# ============================================
# CONFIGURACIÓN DE AUTENTICACIÓN
# ============================================

st.set_page_config(
    page_title="Dashboard de Indicadores de Mantenimiento Mecánico Fortidex",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed"
)

USUARIOS = {
    "w.jimenez@fortidex.com": {
        "nombre": "Wilmer Jimenez",
        "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
        "rol": "administrador"
    },
    "gerencia@grancerv.com": {
        "nombre": "Jairo Granga",
        "password_hash": hashlib.sha256("grancerv123".encode()).hexdigest(),
        "rol": "promotor"
    },
    "tecnico@fortidex.com": {
        "nombre": "Técnico",
        "password_hash": hashlib.sha256("tec123".encode()).hexdigest(),
        "rol": "tecnico"
    }
}

def verificar_login(email, password):
    email = (email or "").strip().lower()
    if email in USUARIOS:
        password_hash = hashlib.sha256((password or "").encode()).hexdigest()
        if USUARIOS[email]["password_hash"] == password_hash:
            return True, USUARIOS[email]
    return False, None

def do_login():
    email = st.session_state.get("login_email", "").strip().lower()
    password = st.session_state.get("login_password", "")

    if not email or not password:
        st.session_state["login_msg"] = ("warning", "⚠️ Por favor, complete todos los campos.")
        return

    login_exitoso, usuario_info = verificar_login(email, password)
    if login_exitoso:
        st.session_state["autenticado"] = True
        st.session_state["usuario"] = usuario_info
        st.session_state["email"] = email
        st.session_state["login_msg"] = ("success", f"✅ ¡Bienvenido, {usuario_info['nombre']}!")
    else:
        st.session_state["login_msg"] = ("error", "❌ Credenciales incorrectas. Intente nuevamente.")

def clear_login():
    st.session_state["login_email"] = ""
    st.session_state["login_password"] = ""
    st.session_state["login_msg"] = None

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if "login_msg" not in st.session_state:
    st.session_state["login_msg"] = None

def mostrar_login():
    st.markdown(
        "<h1 style='text-align: center;'>🔐 Dashboard de Mantenimiento Mecánico Fortidex</h1>",
        unsafe_allow_html=True
    )

    st.markdown("""
    <div style='text-align: center; padding: 20px; background-color: #ffa500; border-radius: 10px;'>
        <h1 style='color: #1f77b4;'>Acceso Restringido</h1>
        <p>Por favor, ingrese sus credenciales para acceder al sistema.</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        with st.container():
            st.markdown("<div style='padding: 30px;'>", unsafe_allow_html=True)

            st.text_input(
                "📧 Correo Electrónico",
                key="login_email",
                placeholder="usuario@fortidex.com",
                on_change=do_login
            )
            st.text_input(
                "🔑 Contraseña",
                key="login_password",
                type="password",
                placeholder="Ingrese su contraseña",
                on_change=do_login
            )

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                login_btn = st.button("🚀 Iniciar Sesión", use_container_width=True)
            with col_btn2:
                clear_btn = st.button("🔄 Limpiar", use_container_width=True, type="secondary")

            if login_btn:
                do_login()

            if clear_btn:
                clear_login()
                st.rerun()

            if st.session_state["login_msg"]:
                tipo, texto = st.session_state["login_msg"]
                if tipo == "success":
                    st.success(texto)
                elif tipo == "error":
                    st.error(texto)
                elif tipo == "warning":
                    st.warning(texto)

            st.markdown("</div>", unsafe_allow_html=True)
           
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 0.9em;'>
        <p>© 2026 Fortidex - Sistema de Gestión de Mantenimiento</p>
        <p>Versión 2.0 | Acceso restringido a personal autorizado</p>
    </div>
    """, unsafe_allow_html=True)

if not st.session_state["autenticado"]:
    mostrar_login()
    st.stop()

# ============================================
# CONTINUACIÓN DEL CÓDIGO ORIGINAL CON MEJORAS
# ============================================

def mostrar_info_usuario():
    with st.sidebar:
        st.markdown("---")
        st.markdown(f"### 👤 {st.session_state['usuario']['nombre']}")
        st.markdown(f"**Rol:** {st.session_state['usuario']['rol'].title()}")
        st.markdown(f"**Email:** {st.session_state['email']}")
        
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            for key in list(st.session_state.keys()):
                if key not in ["data", "personal_data", "last_update"]:
                    del st.session_state[key]
            st.session_state["autenticado"] = False
            st.rerun()

COLOR_PALETTE = {
    'pastel': ['#AEC6CF', '#FFB3BA', '#FFDFBA', '#BAFFC9', '#BAE1FF', '#F0E6EF', '#C9C9FF', '#FFC9F0'],
    'tipo_mtto': {
        'PREVENTIVO': '#87CEEB',
        'BASADO EN CONDICIÓN': '#00008B',
        'CORRECTIVO PROGRAMADO': '#FFD700',
        'CORRECTIVO DE EMERGENCIA': '#FF0000',
        'MEJORA DE SISTEMA': '#32CD32'
    },
    'estado_orden': {
        'CULMINADAS': '#32CD32',
        'EN EJECUCIÓN': '#FFD700',
        'RETRASADAS': '#FFA500',
        'PROYECTADAS': '#52b3f3',
        'TOTAL_PLANIFICADAS': "#02BFF8"
    }
}

@st.cache_data(ttl=300)
def load_overtime_data_from_google_sheets():
    """Carga específicamente datos de DETALLE_HE (horas extras)"""
    try:
        sheet_id = "1X3xgXkeyoei0WkgoNV54zx83XkIKhDlOVEo93lsaFB0"
        gsheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        
        st.info("📥 Cargando datos de HORAS EXTRAS desde hoja 'DETALLE_HE'...")
        
        df = pd.read_excel(gsheet_url, sheet_name='DETALLE_HE')
        
        st.success(f"✅ Datos crudos de DETALLE_HE cargados: {len(df)} registros")
        
        df_clean = clean_overtime_data(df)
        
        if not df_clean.empty:
            st.success(f"✅ Datos de horas extras procesados: {len(df_clean)} registros")
            
            if 'RESPONSABLE' in df_clean.columns:
                tecnicos = df_clean['RESPONSABLE'].nunique()
                st.info(f"   👷 Técnicos únicos (nombres): {tecnicos}")
            
            if 'HORAS_EXTRAS' in df_clean.columns:
                total_horas = df_clean['HORAS_EXTRAS'].sum()
                st.info(f"   ⏰ Horas extras totales: {total_horas:.2f} horas")
            
            if 'SALDO_HORAS_EXTRAS' in df_clean.columns:
                total_saldo = df_clean['SALDO_HORAS_EXTRAS'].sum()
                st.info(f"   💰 Valor total horas extras: ${total_saldo:,.2f}")
        
        return df_clean
    except Exception as e:
        st.error(f"❌ Error cargando DETALLE_HE: {e}")
        return pd.DataFrame()

def verify_overtime_data(overtime_df):
    if overtime_df.empty:
        return "❌ No hay datos cargados"
    
    issues = []
    required_cols = ['INICIO_HORAS_EXTRAS', 'RESPONSABLE', 'HORAS_EXTRAS', 'SALDO_HORAS_EXTRAS']
    missing_cols = [col for col in required_cols if col not in overtime_df.columns]
    
    if missing_cols:
        issues.append(f"Faltan columnas: {missing_cols}")
    
    if 'INICIO_HORAS_EXTRAS' in overtime_df.columns:
        if not pd.api.types.is_datetime64_any_dtype(overtime_df['INICIO_HORAS_EXTRAS']):
            issues.append("INICIO_HORAS_EXTRAS no es tipo datetime")
        else:
            null_dates = overtime_df['INICIO_HORAS_EXTRAS'].isna().sum()
            if null_dates > 0:
                issues.append(f"Hay {null_dates} fechas nulas en INICIO_HORAS_EXTRAS")
    
    if 'HORAS_EXTRAS' in overtime_df.columns:
        zero_hours = (overtime_df['HORAS_EXTRAS'] == 0).sum()
        if zero_hours == len(overtime_df):
            issues.append("Todas las horas extras son 0")
    
    if issues:
        return "⚠️ " + "; ".join(issues)
    else:
        return "✅ Datos OK"

def clean_responsable_column(df):
    if 'RESPONSABLE' in df.columns:
        df['RESPONSABLE'] = df['RESPONSABLE'].astype(str)
        df['RESPONSABLE'] = df['RESPONSABLE'].replace('nan', '').replace('None', '')
        df['RESPONSABLE'] = df['RESPONSABLE'].str.strip()
    return df

def clean_overtime_data(df):
    df_clean = df.copy()
    df_clean.columns = [str(col).strip().upper() for col in df_clean.columns]
    
    if 'RESPONSABLE_N' in df_clean.columns:
        df_clean['RESPONSABLE_NOMBRE'] = df_clean['RESPONSABLE_N'].astype(str).str.strip()
    else:
        df_clean['RESPONSABLE_NOMBRE'] = ''
    
    if 'RESPONSABLE' in df_clean.columns:
        df_clean['RESPONSABLE_CODIGO'] = df_clean['RESPONSABLE'].astype(str).str.strip()
    else:
        df_clean['RESPONSABLE_CODIGO'] = ''
    
    df_clean['RESPONSABLE'] = df_clean['RESPONSABLE_NOMBRE']
    
    if 'HORAS EXTRAS' in df_clean.columns:
        df_clean = df_clean.rename(columns={'HORAS EXTRAS': 'HORAS_EXTRAS'})
    
    if 'SALDO  HORAS EXTRAS' in df_clean.columns:
        df_clean = df_clean.rename(columns={'SALDO  HORAS EXTRAS': 'SALDO_HORAS_EXTRAS'})
    
    if 'INICIO_HORAS_EXTRAS' in df_clean.columns:
        try:
            df_clean['INICIO_HORAS_EXTRAS'] = pd.to_datetime(
                df_clean['INICIO_HORAS_EXTRAS'], 
                errors='coerce',
                dayfirst=True
            )
        except Exception as e:
            df_clean['INICIO_HORAS_EXTRAS'] = pd.NaT
    
    if 'FIN_HORAS_EXTRAS' in df_clean.columns:
        df_clean['FIN_HORAS_EXTRAS'] = pd.to_datetime(
            df_clean['FIN_HORAS_EXTRAS'], 
            errors='coerce',
            dayfirst=True
        )
    
    if 'HORAS_EXTRAS' in df_clean.columns:
        df_clean['HORAS_EXTRAS'] = pd.to_numeric(df_clean['HORAS_EXTRAS'], errors='coerce').fillna(0)
        df_clean['H_EXTRA_MIN'] = df_clean['HORAS_EXTRAS'] * 60
    
    if 'SALDO_HORAS_EXTRAS' in df_clean.columns:
        if df_clean['SALDO_HORAS_EXTRAS'].dtype == 'object':
            df_clean['SALDO_HORAS_EXTRAS'] = df_clean['SALDO_HORAS_EXTRAS'].astype(str)
            df_clean['SALDO_HORAS_EXTRAS'] = df_clean['SALDO_HORAS_EXTRAS'].str.replace('[^\d\.\-]', '', regex=True)
        df_clean['SALDO_HORAS_EXTRAS'] = pd.to_numeric(df_clean['SALDO_HORAS_EXTRAS'], errors='coerce').fillna(0)
    
    if 'OT_ID' in df_clean.columns:
        df_clean['OT'] = df_clean['OT_ID'].astype(str).str.strip()
    
    return df_clean

def calcular_tiempo_disponible(fecha_inicio, fecha_fin):
    num_dias = (fecha_fin - fecha_inicio).days + 1
    tiempo_total = num_dias * 24 * 60
    return tiempo_total, num_dias

def calculate_metrics(df, fecha_inicio, fecha_fin, overtime_data=None):
    if df.empty:
        return {}
    
    m = {}
    m['td'], m['num_dias'] = calcular_tiempo_disponible(fecha_inicio, fecha_fin)
    
    prod_afectada_mask = df['PRODUCCION_AFECTADA'] == 'SI'
    m['tfs'] = df[prod_afectada_mask]['TFS_MIN'].sum() if 'TFS_MIN' in df.columns else 0
    m['tr'] = df[prod_afectada_mask]['TR_MIN'].sum() if 'TR_MIN' in df.columns else 0
    m['tfc'] = df[prod_afectada_mask]['TFC_MIN'].sum() if 'TFC_MIN' in df.columns else 0
    
    m['to'] = max(m['td'] - m['tfs'], 0)
    m['disponibilidad'] = (m['to'] / m['td']) * 100 if m['td'] > 0 else 0
    m['indisponibilidad'] = (m['tfs'] / m['td']) * 100 if m['td'] > 0 else 0
    m['total_fallas'] = len(df[prod_afectada_mask])
    
    m['mtbf'] = m['td'] / m['total_fallas'] if m['total_fallas'] > 0 else 0
    m['mttf'] = m['to'] / m['total_fallas'] if m['total_fallas'] > 0 else 0
    m['mttr'] = m['tr'] / m['total_fallas'] if m['total_fallas'] > 0 else 0
    
    landa = m['total_fallas'] / m['td'] if m['td'] > 0 else 0
    m['mantenibilidad'] = 1 - np.exp(-landa * m['td']) if landa > 0 else 0
    
    tipo_mtto_totals = df.groupby('TIPO DE MTTO')['TR_MIN'].sum()
    total_mtto = tipo_mtto_totals.sum()
    
    if total_mtto > 0:
        m['mp_pct'] = (tipo_mtto_totals.get('PREVENTIVO', 0) / total_mtto) * 100
        m['mbc_pct'] = (tipo_mtto_totals.get('BASADO EN CONDICIÓN', 0) / total_mtto) * 100
        m['mce_pct'] = (tipo_mtto_totals.get('CORRECTIVO DE EMERGENCIA', 0) / total_mtto) * 100
        m['mcp_pct'] = (tipo_mtto_totals.get('CORRECTIVO PROGRAMADO', 0) / total_mtto) * 100
        m['mms_pct'] = (tipo_mtto_totals.get('MEJORA DE SISTEMA', 0) / total_mtto) * 100
    else:
        m['mp_pct'] = m['mbc_pct'] = m['mce_pct'] = m['mcp_pct'] = m['mms_pct'] = 0
    
    if overtime_data is not None and not overtime_data.empty:
        m['horas_extras_acumuladas'] = overtime_data['H_EXTRA_MIN'].sum() if 'H_EXTRA_MIN' in overtime_data.columns else 0
    else:
        m['horas_extras_acumuladas'] = 0
    
    return m

def calculate_reliability_metrics(df, fecha_inicio, fecha_fin):
    if df.empty:
        return {}
    
    emergency_mask = df['TIPO DE MTTO'] == 'CORRECTIVO DE EMERGENCIA'
    df_emergency = df[emergency_mask].copy()
    
    if df_emergency.empty:
        return {}
    
    m = {}
    m['td'], m['num_dias'] = calcular_tiempo_disponible(fecha_inicio, fecha_fin)
    
    m['tr_emergency'] = df_emergency['TR_MIN'].sum() if 'TR_MIN' in df_emergency.columns else 0
    m['tfc_emergency'] = df_emergency['TFC_MIN'].sum() if 'TFC_MIN' in df_emergency.columns else 0
    m['tfs_emergency'] = df_emergency['TFS_MIN'].sum() if 'TFS_MIN' in df_emergency.columns else 0
    m['total_fallas_emergency'] = len(df_emergency)
    m['total_fallas_emergency_con_parada'] = len(df_emergency[df_emergency['PRODUCCION_AFECTADA'] == 'SI'])
    
    if m['total_fallas_emergency'] > 0:
        m['mtbf_emergency'] = m['td'] / m['total_fallas_emergency'] if m['td'] > 0 else 0
        m['mttr_emergency'] = m['tr_emergency'] / m['total_fallas_emergency'] if m['total_fallas_emergency'] > 0 else 0
        
        emergency_prod_mask = (df_emergency['PRODUCCION_AFECTADA'] == 'SI')
        tfs_emergency_prod = df_emergency[emergency_prod_mask]['TFS_MIN'].sum() if 'TFS_MIN' in df_emergency.columns else 0
        to_emergency = max(m['td'] - tfs_emergency_prod, 0)
        m['mttf_emergency'] = to_emergency / m['total_fallas_emergency'] if m['total_fallas_emergency'] > 0 else 0
    else:
        m['mtbf_emergency'] = 0
        m['mttr_emergency'] = 0
        m['mttf_emergency'] = 0
    
    landa_emergency = m['total_fallas_emergency'] / m['td'] if m['td'] > 0 else 0
    m['mantenibilidad_emergency'] = 1 - np.exp(-landa_emergency * m['td']) if landa_emergency > 0 else 0
    m['mantenibilidad_pct'] = m['mantenibilidad_emergency'] * 100
    
    return m

def get_weekly_data(df, fecha_inicio, fecha_fin):
    if df.empty or 'FECHA_DE_INICIO' not in df.columns:
        return pd.DataFrame()
    
    df_weekly = df.copy()
    df_weekly['SEMANA'] = df_weekly['FECHA_DE_INICIO'].dt.isocalendar().week
    df_weekly['AÑO'] = df_weekly['FECHA_DE_INICIO'].dt.year
    df_weekly['SEMANA_STR'] = df_weekly.apply(
        lambda x: f"{x['AÑO']}-S{x['SEMANA']:02d}", axis=1
    )
    
    tiempo_disponible_semanal = {}
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        week_end = min(current_date + timedelta(days=6), fecha_fin)
        days_in_week = (week_end - current_date).days + 1
        minutos_semana = days_in_week * 24 * 60
        semana_num = datetime.combine(current_date, datetime.min.time()).isocalendar().week
        año_num = current_date.year
        semana_str = f"{año_num}-S{semana_num:02d}"
        tiempo_disponible_semanal[semana_str] = minutos_semana
        current_date = week_end + timedelta(days=1)
    
    weekly_data = df_weekly[df_weekly['PRODUCCION_AFECTADA'] == 'SI'].groupby(['SEMANA_STR', 'AÑO', 'SEMANA']).agg({
        'TFS_MIN': 'sum',
        'TR_MIN': 'sum',
        'TFC_MIN': 'sum',
        'PRODUCCION_AFECTADA': lambda x: (x == 'SI').sum()
    }).reset_index()
    
    weekly_data['TDISPONIBLE'] = weekly_data['SEMANA_STR'].map(tiempo_disponible_semanal).fillna(0)
    weekly_data['DISPO_SEMANAL'] = ((weekly_data['TDISPONIBLE'] - weekly_data['TFS_MIN']) / weekly_data['TDISPONIBLE']) * 100
    weekly_data['SEMANA_NUM'] = weekly_data['AÑO'] * 100 + weekly_data['SEMANA']
    weekly_data = weekly_data.sort_values('SEMANA_NUM')
    
    return weekly_data

def crear_velocimetro_mejorado(valor, titulo, valor_min=0, valor_max=100, color_verde=80, color_amarillo=60):
    if valor >= color_verde:
        color_aguja = '#32CD32'
    elif valor >= color_amarillo:
        color_aguja = '#FFD700'
    else:
        color_aguja = '#FF0000'
    
    rango = valor_max - valor_min
    angulo = 180 * (valor - valor_min) / rango if rango > 0 else 90
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatterpolar(
        r=[0.5, 0.5, 0.5],
        theta=[0, 90, 180],
        mode='lines',
        line_color='lightgray',
        line_width=2,
        showlegend=False
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=[0.7, 0.7, 0.7, 0.7],
        theta=[180, 180 * color_amarillo/valor_max, 180 * color_verde/valor_max, 0],
        fill='toself',
        fillcolor='#FF0000',
        line_color='#FF0000',
        opacity=0.3,
        name='Crítico'
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=[0.7, 0.7, 0.7],
        theta=[180 * color_amarillo/valor_max, 180 * color_verde/valor_max, 0],
        fill='toself',
        fillcolor='#FFD700',
        line_color='#FFD700',
        opacity=0.3,
        name='Regular'
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=[0.7, 0.7, 0.7],
        theta=[180 * color_verde/valor_max, 0, 0],
        fill='toself',
        fillcolor='#32CD32',
        line_color='#32CD32',
        opacity=0.3,
        name='Excelente'
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=[0, 0.85, 0],
        theta=[angulo - 5, angulo, angulo + 5],
        mode='lines',
        line_color=color_aguja,
        line_width=4,
        fill='toself',
        fillcolor=color_aguja,
        name='Valor actual'
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=[0.1],
        theta=[angulo],
        mode='markers',
        marker=dict(color='black', size=8),
        showlegend=False
    ))
    
    fig.update_layout(
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 1], showticklabels=False),
            angularaxis=dict(
                rotation=90,
                direction="clockwise",
                tickmode='array',
                tickvals=[0, 45, 90, 135, 180],
                ticktext=[f'{valor_max}', f'{valor_max*0.75}', f'{valor_max*0.5}', f'{valor_max*0.25}', f'{valor_min}'],
                showticklabels=True,
                tickfont=dict(size=14)
            ),
            bgcolor='white'
        ),
        showlegend=False,
        height=350,
        title=dict(text=titulo, font=dict(size=18, color='black'), x=0.5, y=0.95),
        margin=dict(t=100, b=50, l=50, r=50)
    )
    
    fig.add_annotation(
        x=0.5, y=0.5,
        text=f"<b>{valor:.1f}</b>",
        showarrow=False,
        font=dict(size=32, color='black'),
        xref="paper", yref="paper"
    )
    
    if valor_max == 100:
        fig.add_annotation(
            x=0.5, y=0.4,
            text="%",
            showarrow=False,
            font=dict(size=20, color='gray'),
            xref="paper", yref="paper"
        )
    
    return fig

def separar_tecnicos(df):
    if df.empty or 'RESPONSABLE' not in df.columns:
        return df
    
    df_separado = df.copy()
    filas_separadas = []
    delimitadores = [',', ';', '|', '/', '\\', 'y', 'Y', '&']
    
    for idx, row in df_separado.iterrows():
        responsable = str(row['RESPONSABLE']).strip()
        
        if not responsable or responsable.lower() == 'nan':
            filas_separadas.append(row)
            continue
        
        tecnicos_encontrados = []
        encontrado_delimitador = False
        
        for delim in delimitadores:
            if delim in responsable:
                partes = [p.strip() for p in responsable.split(delim) if p.strip()]
                if len(partes) > 1:
                    tecnicos_encontrados.extend(partes)
                    encontrado_delimitador = True
                    break
        
        if not encontrado_delimitador:
            patrones = [r'(\w+\s+\d+\s*,\s*\w+\s+\d+)', r'(\w+\s+y\s+\w+)']
            for patron in patrones:
                coincidencias = re.findall(patron, responsable)
                if coincidencias:
                    if ',' in responsable:
                        tecnicos_encontrados = [t.strip() for t in responsable.split(',') if t.strip()]
                    elif 'y' in responsable.lower():
                        partes = re.split(r'\s+y\s+', responsable, flags=re.IGNORECASE)
                        tecnicos_encontrados = [p.strip() for p in partes if p.strip()]
                    encontrado_delimitador = True
                    break
        
        if len(tecnicos_encontrados) > 1:
            for tecnico in tecnicos_encontrados:
                nueva_fila = row.copy()
                nueva_fila['RESPONSABLE'] = tecnico
                if 'TR_MIN' in nueva_fila:
                    nueva_fila['TR_MIN'] = row['TR_MIN'] if pd.notna(row['TR_MIN']) else 0
                if 'H_EXTRA_MIN' in nueva_fila:
                    nueva_fila['H_EXTRA_MIN'] = row['H_EXTRA_MIN'] if pd.notna(row['H_EXTRA_MIN']) else 0
                if 'H_NORMAL_MIN' in nueva_fila:
                    nueva_fila['H_NORMAL_MIN'] = row['H_NORMAL_MIN'] if pd.notna(row['H_NORMAL_MIN']) else 0
                filas_separadas.append(nueva_fila)
        else:
            filas_separadas.append(row)
    
    df_resultado = pd.DataFrame(filas_separadas)
    df_resultado['RESPONSABLE'] = df_resultado['RESPONSABLE'].astype(str).str.strip()
    return df_resultado

@st.cache_data(ttl=300)
def load_personal_data_from_google_sheets():
    try:
        sheet_id = "1X3xgXkeyoei0WkgoNV54zx83XkIKhDlOVEo93lsaFB0"
        gsheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        df_personal = pd.read_excel(gsheet_url, sheet_name='PERSONAL')
        df_personal.columns = [col.strip().upper() for col in df_personal.columns]
        return df_personal
    except Exception as e:
        st.error(f"Error al cargar datos del personal desde Google Sheets: {e}")
        return pd.DataFrame()

def calculate_overtime_costs_from_details(overtime_data, personal_data):
    if overtime_data.empty:
        return pd.DataFrame(), pd.DataFrame(), "No hay datos de horas extras en el período seleccionado"
    
    required_columns = ['RESPONSABLE', 'H_EXTRA_MIN', 'SALDO_HORAS_EXTRAS', 'INICIO_HORAS_EXTRAS']
    missing_columns = [col for col in required_columns if col not in overtime_data.columns]
    
    if missing_columns:
        st.warning(f"Faltan columnas en datos de horas extras: {missing_columns}")
        return pd.DataFrame(), pd.DataFrame(), f"Faltan columnas: {missing_columns}"
    
    try:
        df_costs = overtime_data.copy()
        df_costs = df_costs[df_costs['RESPONSABLE'].notna()]
        df_costs['RESPONSABLE'] = df_costs['RESPONSABLE'].astype(str).str.strip()
        
        if df_costs.empty:
            return pd.DataFrame(), pd.DataFrame(), "No hay registros con responsable asignado"
        
        df_costs['H_EXTRA_HORAS'] = df_costs['H_EXTRA_MIN'] / 60
        
        if not pd.api.types.is_datetime64_any_dtype(df_costs['INICIO_HORAS_EXTRAS']):
            df_costs['INICIO_HORAS_EXTRAS'] = pd.to_datetime(df_costs['INICIO_HORAS_EXTRAS'], errors='coerce')
        
        df_costs = df_costs[df_costs['INICIO_HORAS_EXTRAS'].notna()]
        
        if df_costs.empty:
            return pd.DataFrame(), pd.DataFrame(), "No hay fechas válidas en los datos"
        
        df_costs['SEMANA'] = df_costs['INICIO_HORAS_EXTRAS'].dt.isocalendar().week
        df_costs['AÑO'] = df_costs['INICIO_HORAS_EXTRAS'].dt.year
        df_costs['SEMANA_STR'] = df_costs['AÑO'].astype(str) + '-S' + df_costs['SEMANA'].astype(str).str.zfill(2)
        df_costs['SEMANA_STR'] = df_costs['SEMANA_STR'].astype(str)
        
        weekly_costs = df_costs.groupby(['SEMANA_STR', 'RESPONSABLE']).agg({
            'SALDO_HORAS_EXTRAS': 'sum',
            'H_EXTRA_HORAS': 'sum',
            'H_EXTRA_MIN': 'sum'
        }).reset_index()
        weekly_costs = weekly_costs.rename(columns={'RESPONSABLE': 'TECNICO', 'SALDO_HORAS_EXTRAS': 'COSTO_TOTAL'})
        
        accumulated_costs = df_costs.groupby('RESPONSABLE').agg({
            'SALDO_HORAS_EXTRAS': 'sum',
            'H_EXTRA_HORAS': 'sum',
            'H_EXTRA_MIN': 'sum'
        }).reset_index()
        accumulated_costs = accumulated_costs.rename(columns={'RESPONSABLE': 'TECNICO', 'SALDO_HORAS_EXTRAS': 'COSTO_TOTAL'})
        accumulated_costs = accumulated_costs.sort_values('COSTO_TOTAL', ascending=False)
        
        total_costo = accumulated_costs['COSTO_TOTAL'].sum()
        total_horas = accumulated_costs['H_EXTRA_HORAS'].sum()
        costo_promedio_hora = total_costo / total_horas if total_horas > 0 else 0
        
        mensaje = f"Cálculo exitoso | Costo total: ${total_costo:,.2f} | Horas totales: {total_horas:,.1f} | Costo promedio/hora: ${costo_promedio_hora:,.2f}"
        
        return weekly_costs, accumulated_costs, mensaje
    except Exception as e:
        st.error(f"Error al calcular costos de horas extras: {e}")
        return pd.DataFrame(), pd.DataFrame(), f"Error en cálculo: {str(e)}"

def show_detailed_costs_info(weekly_costs, accumulated_costs, personal_data):
    st.subheader("📋 Información Detallada de Costos")
    
    if accumulated_costs.empty:
        st.info("No hay datos de costos acumulados para mostrar.")
        return
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_costo = accumulated_costs['COSTO_TOTAL'].sum()
        st.metric("Costo Total Horas Extras", f"${total_costo:,.2f}")
    
    with col2:
        total_horas = accumulated_costs['H_EXTRA_HORAS'].sum()
        st.metric("Horas Extras Totales", f"{total_horas:,.1f} horas")
    
    with col3:
        costo_promedio = total_costo / total_horas if total_horas > 0 else 0
        st.metric("Costo Promedio por Hora", f"${costo_promedio:,.2f}")
    
    with col4:
        num_tecnicos = len(accumulated_costs)
        st.metric("Técnicos con Horas Extras", f"{num_tecnicos}")
    
    st.subheader("📊 Detalle de Costos por Técnico")
    
    tabla_detalle = accumulated_costs.copy()
    tabla_detalle['COSTO_TOTAL_FMT'] = tabla_detalle['COSTO_TOTAL'].apply(lambda x: f"${x:,.2f}")
    tabla_detalle['HORAS_EXTRA_FMT'] = tabla_detalle['H_EXTRA_HORAS'].apply(lambda x: f"{x:,.2f}")
    tabla_detalle['COSTO_POR_HORA'] = tabla_detalle.apply(
        lambda x: x['COSTO_TOTAL'] / x['H_EXTRA_HORAS'] if x['H_EXTRA_HORAS'] > 0 else 0, axis=1
    )
    tabla_detalle['COSTO_POR_HORA_FMT'] = tabla_detalle['COSTO_POR_HORA'].apply(lambda x: f"${x:,.2f}")
    tabla_detalle['PORCENTAJE'] = (tabla_detalle['COSTO_TOTAL'] / total_costo * 100) if total_costo > 0 else 0
    tabla_detalle['PORCENTAJE_FMT'] = tabla_detalle['PORCENTAJE'].apply(lambda x: f"{x:.1f}%")
    
    columnas_mostrar = ['TECNICO', 'HORAS_EXTRA_FMT', 'COSTO_POR_HORA_FMT', 'COSTO_TOTAL_FMT', 'PORCENTAJE_FMT']
    tabla_detalle = tabla_detalle[columnas_mostrar]
    tabla_detalle.columns = ['Técnico', 'Horas Extras', 'Costo por Hora', 'Costo Total', '% del Total']
    
    st.dataframe(tabla_detalle, use_container_width=True)
    
    if not weekly_costs.empty:
        with st.expander("Ver datos semanales detallados"):
            weekly_formatted = weekly_costs.copy()
            weekly_formatted['COSTO_TOTAL_FMT'] = weekly_formatted['COSTO_TOTAL'].apply(lambda x: f"${x:,.2f}")
            weekly_formatted['HORAS_EXTRA_FMT'] = weekly_formatted['H_EXTRA_HORAS'].apply(lambda x: f"{x:,.2f}")
            st.dataframe(
                weekly_formatted[['SEMANA_STR', 'TECNICO', 'HORAS_EXTRA_FMT', 'COSTO_TOTAL_FMT']],
                use_container_width=True
            )

def calcular_duracion_minutos(fecha_inicio, hora_inicio, fecha_fin, hora_fin):
    try:
        datetime_inicio = pd.to_datetime(fecha_inicio.strftime('%Y-%m-%d') + ' ' + str(hora_inicio))
        datetime_fin = pd.to_datetime(fecha_fin.strftime('%Y-%m-%d') + ' ' + str(hora_fin))
        duracion = (datetime_fin - datetime_inicio).total_seconds() / 60
        return max(duracion, 0)
    except:
        return 0

@st.cache_data(ttl=300)
def load_data_from_google_sheets():
    try:
        sheet_id = "1X3xgXkeyoei0WkgoNV54zx83XkIKhDlOVEo93lsaFB0"
        gsheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        df = pd.read_excel(gsheet_url, sheet_name='DATAMTTO')
        df = clean_and_prepare_data(df)
        return df
    except Exception as e:
        st.error(f"Error al cargar datos desde Google Sheets: {e}")
        st.info("Asegúrate de que el archivo de Google Sheets sea público y accesible")
        return pd.DataFrame()

def clean_and_prepare_data(df):
    df_clean = df.copy()
    
    df_clean = df_clean.rename(columns={
        'FECHA DE INICIO': 'FECHA_DE_INICIO',
        'FECHA DE FIN': 'FECHA_DE_FIN',
        'Tiempo Prog (min)': 'TIEMPO_PROG_MIN',
        'PRODUCCIÓN AFECTADA (SI-NO)': 'PRODUCCION_AFECTADA',
        'TIEMPO ESTIMADO DIARIO (min)': 'TDISPONIBLE_OLD',
        'TR (min)': 'TR_MIN',
        'TFC (min)': 'TFC_MIN',
        'TFS (min)': 'TFS_MIN',
        'h normal (min)': 'H_NORMAL_MIN',
        'h extra (min)': 'H_EXTRA_MIN',
        'HORA PARADA DE MÁQUINA': 'HORA_PARADA',
        'HORA INICIO': 'HORA_INICIO',
        'HORA FINAL': 'HORA_FINAL',
        'HORA DE ARRANQUE': 'HORA_ARRANQUE'
    })
    
    if 'UBICACIÓN TÉCNICA NOMBRE' in df_clean.columns:
        df_clean['UBICACIÓN TÉCNICA'] = df_clean['UBICACIÓN TÉCNICA NOMBRE']
    elif 'UBICACIÓN TÉCNICA' not in df_clean.columns and 'UBICACION TECNICA' in df_clean.columns:
        df_clean = df_clean.rename(columns={'UBICACION TECNICA': 'UBICACIÓN TÉCNICA'})
    elif 'UBICACIÓN TÉCNica' not in df_clean.columns and 'Ubicación Técnica' in df_clean.columns:
        df_clean = df_clean.rename(columns={'Ubicación Técnica': 'UBICACIÓN TÉCNICA'})
    
    if 'EQUIPO NOMBRE' in df_clean.columns:
        df_clean['EQUIPO'] = df_clean['EQUIPO NOMBRE']
    
    if 'CONJUNTO NOMBRE' in df_clean.columns:
        df_clean['CONJUNTO'] = df_clean['CONJUNTO NOMBRE']
    
    if 'RESPONSABLE NOMBRE' in df_clean.columns:
        df_clean['RESPONSABLE'] = df_clean['RESPONSABLE NOMBRE']
    
    df_clean['FECHA_DE_INICIO'] = pd.to_datetime(df_clean['FECHA_DE_INICIO'])
    df_clean['FECHA_DE_FIN'] = pd.to_datetime(df_clean['FECHA_DE_FIN'])
    
    df_clean['TR_MIN_CALCULADO'] = df_clean.apply(
        lambda x: calcular_duracion_minutos(
            x['FECHA_DE_INICIO'], x['HORA_INICIO'], 
            x['FECHA_DE_FIN'], x['HORA_FINAL']
        ), axis=1
    )
    
    if 'TR_MIN' in df_clean.columns:
        df_clean['TR_MIN'] = df_clean.apply(
            lambda x: x['TR_MIN_CALCULADO'] if pd.isna(x['TR_MIN']) or x['TR_MIN'] == 0 else x['TR_MIN'], axis=1
        )
    else:
        df_clean['TR_MIN'] = df_clean['TR_MIN_CALCULADO']
    
    numeric_columns = ['TR_MIN', 'TFC_MIN', 'TFS_MIN', 'TDISPONIBLE_OLD', 'TIEMPO_PROG_MIN']
    for col in numeric_columns:
        if col in df_clean.columns:
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)
    
    if 'H_EXTRA_MIN' in df_clean.columns:
        df_clean = df_clean.drop(columns=['H_EXTRA_MIN'])
    
    if 'RESPONSABLE' in df_clean.columns:
        df_clean['RESPONSABLE'] = df_clean['RESPONSABLE'].astype(str).str.strip()
    
    return df_clean

def get_weekly_overtime_data(overtime_data):
    if overtime_data.empty:
        return pd.DataFrame()
    
    required_cols = ['INICIO_HORAS_EXTRAS', 'RESPONSABLE', 'H_EXTRA_MIN', 'SALDO_HORAS_EXTRAS']
    for col in required_cols:
        if col not in overtime_data.columns:
            st.warning(f"Falta columna {col} en datos de horas extras")
            return pd.DataFrame()
    
    df_weekly = overtime_data.copy()
    
    if not pd.api.types.is_datetime64_any_dtype(df_weekly['INICIO_HORAS_EXTRAS']):
        df_weekly['INICIO_HORAS_EXTRAS'] = pd.to_datetime(df_weekly['INICIO_HORAS_EXTRAS'], errors='coerce')
    
    df_weekly = df_weekly[df_weekly['INICIO_HORAS_EXTRAS'].notna()]
    
    if df_weekly.empty:
        return pd.DataFrame()
    
    df_weekly['RESPONSABLE'] = df_weekly['RESPONSABLE'].astype(str).str.strip()
    
    try:
        df_weekly['SEMANA'] = df_weekly['INICIO_HORAS_EXTRAS'].dt.isocalendar().week
        df_weekly['AÑO'] = df_weekly['INICIO_HORAS_EXTRAS'].dt.year
        df_weekly['SEMANA_STR'] = df_weekly.apply(
            lambda x: f"{int(x['AÑO'])}-S{int(x['SEMANA']):02d}", axis=1
        )
    except Exception as e:
        st.error(f"Error al calcular semana: {e}")
        return pd.DataFrame()
    
    try:
        weekly_overtime = df_weekly.groupby(['SEMANA_STR', 'AÑO', 'SEMANA', 'RESPONSABLE']).agg({
            'H_EXTRA_MIN': 'sum',
            'SALDO_HORAS_EXTRAS': 'sum'
        }).reset_index()
        
        weekly_overtime['H_EXTRA_HORAS'] = weekly_overtime['H_EXTRA_MIN'] / 60
        weekly_overtime['SEMANA_NUM'] = weekly_overtime['AÑO'] * 100 + weekly_overtime['SEMANA']
        weekly_overtime = weekly_overtime.sort_values('SEMANA_NUM')
        
        return weekly_overtime
    except Exception as e:
        st.error(f"Error al agrupar datos de horas extras: {e}")
        return pd.DataFrame()

def get_accumulated_overtime_data(overtime_data):
    if overtime_data.empty or 'RESPONSABLE' not in overtime_data.columns:
        return pd.DataFrame()
    
    overtime_data['RESPONSABLE'] = overtime_data['RESPONSABLE'].astype(str).str.strip()
    
    overtime_tech_data = overtime_data.groupby('RESPONSABLE').agg({
        'H_EXTRA_MIN': 'sum',
        'SALDO_HORAS_EXTRAS': 'sum'
    }).reset_index()
    
    overtime_tech_data['H_EXTRA_HORAS'] = overtime_tech_data['H_EXTRA_MIN'] / 60
    overtime_tech_data = overtime_tech_data.sort_values('H_EXTRA_HORAS', ascending=False)
    
    return overtime_tech_data

def get_weekly_technician_hours(df, overtime_data):
    if df.empty or 'FECHA_DE_INICIO' not in df.columns or 'RESPONSABLE' not in df.columns:
        return pd.DataFrame()
    
    df_separado = separar_tecnicos(df)
    df_weekly = df_separado.copy()
    
    df_weekly['SEMANA'] = df_weekly['FECHA_DE_INICIO'].dt.isocalendar().week
    df_weekly['AÑO'] = df_weekly['FECHA_DE_INICIO'].dt.year
    df_weekly['SEMANA_STR'] = df_weekly.apply(lambda x: f"{x['AÑO']}-S{x['SEMANA']:02d}", axis=1)
    df_weekly['RESPONSABLE'] = df_weekly['RESPONSABLE'].astype(str).str.strip()
    df_weekly['SEMANA_STR'] = df_weekly['SEMANA_STR'].astype(str)
    
    weekly_normal_hours = df_weekly.groupby(['SEMANA_STR', 'AÑO', 'SEMANA', 'RESPONSABLE']).agg({
        'TR_MIN': 'sum'
    }).reset_index()
    weekly_normal_hours['TR_HORAS'] = weekly_normal_hours['TR_MIN'] / 60
    
    weekly_overtime_hours = get_weekly_overtime_data(overtime_data)
    
    if not weekly_overtime_hours.empty:
        weekly_overtime_hours['RESPONSABLE'] = weekly_overtime_hours['RESPONSABLE'].astype(str).str.strip()
        weekly_overtime_hours['SEMANA_STR'] = weekly_overtime_hours['SEMANA_STR'].astype(str)
    
    if not weekly_overtime_hours.empty:
        weekly_tech_data = pd.merge(
            weekly_normal_hours,
            weekly_overtime_hours[['SEMANA_STR', 'RESPONSABLE', 'H_EXTRA_MIN', 'H_EXTRA_HORAS']],
            on=['SEMANA_STR', 'RESPONSABLE'],
            how='outer'
        ).fillna(0)
    else:
        weekly_tech_data = weekly_normal_hours.copy()
        weekly_tech_data['H_EXTRA_MIN'] = 0
        weekly_tech_data['H_EXTRA_HORAS'] = 0
    
    weekly_tech_data['SEMANA_NUM'] = weekly_tech_data['AÑO'] * 100 + weekly_tech_data['SEMANA']
    weekly_tech_data = weekly_tech_data.sort_values('SEMANA_NUM')
    
    return weekly_tech_data

def get_accumulated_technician_hours(df, overtime_data):
    if df.empty or 'RESPONSABLE' not in df.columns:
        return pd.DataFrame()
    
    df_separado = separar_tecnicos(df)
    
    normal_tech_data = df_separado.groupby('RESPONSABLE').agg({'TR_MIN': 'sum'}).reset_index()
    normal_tech_data['TR_HORAS'] = normal_tech_data['TR_MIN'] / 60
    
    overtime_tech_data = get_accumulated_overtime_data(overtime_data)
    normal_tech_data['RESPONSABLE'] = normal_tech_data['RESPONSABLE'].astype(str).str.strip()
    
    if not overtime_tech_data.empty:
        overtime_tech_data['RESPONSABLE'] = overtime_tech_data['RESPONSABLE'].astype(str).str.strip()
        tech_data = pd.merge(
            normal_tech_data,
            overtime_tech_data[['RESPONSABLE', 'H_EXTRA_MIN', 'H_EXTRA_HORAS', 'SALDO_HORAS_EXTRAS']],
            on='RESPONSABLE',
            how='outer'
        ).fillna(0)
    else:
        tech_data = normal_tech_data.copy()
        tech_data['H_EXTRA_MIN'] = 0
        tech_data['H_EXTRA_HORAS'] = 0
        tech_data['SALDO_HORAS_EXTRAS'] = 0
    
    tech_data = tech_data.sort_values('TR_HORAS', ascending=False)
    return tech_data

def get_weekly_emergency_data(df):
    if df.empty or 'FECHA_DE_INICIO' not in df.columns:
        return pd.DataFrame()
    
    df_weekly = df.copy()
    df_weekly['SEMANA'] = df_weekly['FECHA_DE_INICIO'].dt.isocalendar().week
    df_weekly['AÑO'] = df_weekly['FECHA_DE_INICIO'].dt.year
    df_weekly['SEMANA_STR'] = df_weekly.apply(lambda x: f"{x['AÑO']}-S{x['SEMANA']:02d}", axis=1)
    
    df_emergency = df_weekly[df_weekly['TIPO DE MTTO'] == 'CORRECTIVO DE EMERGENCIA'].copy()
    
    if df_emergency.empty:
        return pd.DataFrame()
    
    weekly_emergency_data = df_emergency.groupby(['SEMANA_STR', 'AÑO', 'SEMANA']).agg({
        'TR_MIN': 'sum',
        'TFC_MIN': 'sum',
        'TFS_MIN': 'sum'
    }).reset_index()
    
    weekly_emergency_counts = df_emergency.groupby(['SEMANA_STR', 'AÑO', 'SEMANA']).size().reset_index(name='NUM_ORDENES_EMERGENCIA')
    weekly_emergency_parada_counts = df_emergency[df_emergency['PRODUCCION_AFECTADA'] == 'SI'].groupby(['SEMANA_STR', 'AÑO', 'SEMANA']).size().reset_index(name='NUM_ORDENES_EMERGENCIA_PARADA')
    
    weekly_emergency_data = weekly_emergency_data.merge(weekly_emergency_counts, on=['SEMANA_STR', 'AÑO', 'SEMANA'], how='left')
    weekly_emergency_data = weekly_emergency_data.merge(weekly_emergency_parada_counts, on=['SEMANA_STR', 'AÑO', 'SEMANA'], how='left')
    weekly_emergency_data['NUM_ORDENES_EMERGENCIA_PARADA'] = weekly_emergency_data['NUM_ORDENES_EMERGENCIA_PARADA'].fillna(0)
    
    weekly_emergency_data['MTTR_SEMANAL'] = weekly_emergency_data.apply(
        lambda row: row['TR_MIN'] / row['NUM_ORDENES_EMERGENCIA'] if row['NUM_ORDENES_EMERGENCIA'] > 0 else 0, axis=1
    )
    
    weekly_emergency_data['SEMANA_NUM'] = weekly_emergency_data['AÑO'] * 100 + weekly_emergency_data['SEMANA']
    weekly_emergency_data = weekly_emergency_data.sort_values('SEMANA_NUM')
    
    return weekly_emergency_data

def get_monthly_plan_data(df, year=2026):
    meses_todos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), (5, 'Mayo'), (6, 'Junio'),
        (7, 'Julio'), (8, 'Agosto'), (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    monthly_data = pd.DataFrame(meses_todos, columns=['MES', 'MES_NOMBRE'])
    monthly_data['AÑO'] = year
    monthly_data['MES_ORDEN'] = monthly_data['MES']
    monthly_data['TOTAL_PLANIFICADAS'] = 0
    monthly_data['ORDENES_CULMINADAS'] = 0
    monthly_data['ORDENES_EN_EJECUCION'] = 0
    monthly_data['ORDENES_RETRASADAS'] = 0
    monthly_data['ORDENES_PROYECTADAS'] = 0
    monthly_data['CUMPLIMIENTO_PCT'] = 0
    
    if df.empty or 'FECHA_DE_INICIO' not in df.columns or 'TIPO DE MTTO' not in df.columns:
        return monthly_data
    
    tipos_planificados = ['PREVENTIVO', 'BASADO EN CONDICIÓN', 'MEJORA DE SISTEMA']
    df_plan = df[df['TIPO DE MTTO'].isin(tipos_planificados)].copy()
    df_plan = df_plan[df_plan['FECHA_DE_INICIO'].dt.year == year]
    
    if df_plan.empty:
        return monthly_data
    
    fecha_actual = datetime.now().date()
    fecha_corte = fecha_actual - timedelta(days=1)
    
    df_plan['FECHA_INICIO_DATE'] = df_plan['FECHA_DE_INICIO'].dt.date
    df_plan['FECHA_FIN_DATE'] = df_plan['FECHA_DE_FIN'].dt.date
    
    df_plan_filtrado = df_plan[
        (df_plan['FECHA_INICIO_DATE'] <= fecha_corte) & 
        (df_plan['FECHA_FIN_DATE'] <= fecha_corte)
    ].copy()
    
    if df_plan_filtrado.empty:
        return monthly_data
    
    df_plan_filtrado['MES'] = df_plan_filtrado['FECHA_DE_INICIO'].dt.month
    df_plan_filtrado['MES_NOMBRE'] = df_plan_filtrado['MES'].map(dict(meses_todos))
    df_plan_filtrado['AÑO'] = df_plan_filtrado['FECHA_DE_INICIO'].dt.year
    
    if 'STATUS' not in df_plan_filtrado.columns:
        df_plan_filtrado['STATUS_NORM'] = 'CULMINADO'
    else:
        df_plan_filtrado['STATUS_NORM'] = df_plan_filtrado['STATUS'].astype(str).str.upper().str.strip()
        df_plan_filtrado.loc[df_plan_filtrado['STATUS_NORM'].str.contains('CULMINAD'), 'STATUS_NORM'] = 'CULMINADO'
        df_plan_filtrado.loc[df_plan_filtrado['STATUS_NORM'].str.contains('PROCESO') | 
                           df_plan_filtrado['STATUS_NORM'].str.contains('PROGRESO') |
                           df_plan_filtrado['STATUS_NORM'].str.contains('EJECUCI'), 'STATUS_NORM'] = 'EN PROCESO'
        df_plan_filtrado.loc[df_plan_filtrado['STATUS_NORM'].str.contains('PENDIENTE'), 'STATUS_NORM'] = 'PENDIENTE'
    
    mask_culminadas = df_plan_filtrado['STATUS_NORM'] == 'CULMINADO'
    mask_en_ejecucion = df_plan_filtrado['STATUS_NORM'] == 'EN PROCESO'
    
    mask_proyectadas = (df_plan_filtrado['STATUS_NORM'] == 'PENDIENTE') & (
        (df_plan_filtrado['FECHA_INICIO_DATE'] >= fecha_actual) |
        ((df_plan_filtrado['FECHA_INICIO_DATE'] < fecha_actual) & (df_plan_filtrado['FECHA_FIN_DATE'] >= fecha_actual))
    )
    
    mask_retrasadas = (df_plan_filtrado['STATUS_NORM'] == 'PENDIENTE') & (
        (df_plan_filtrado['FECHA_INICIO_DATE'] < fecha_actual) & (df_plan_filtrado['FECHA_FIN_DATE'] < fecha_actual)
    )
    
    monthly_real_data = df_plan_filtrado.groupby(['AÑO', 'MES', 'MES_NOMBRE']).agg({'TIPO DE MTTO': 'count'}).reset_index()
    monthly_real_data = monthly_real_data.rename(columns={'TIPO DE MTTO': 'TOTAL_PLANIFICADAS'})
    
    df_culminadas = df_plan_filtrado[mask_culminadas]
    monthly_culminadas = df_culminadas.groupby(['AÑO', 'MES', 'MES_NOMBRE']).agg({'TIPO DE MTTO': 'count'}).reset_index()
    monthly_culminadas = monthly_culminadas.rename(columns={'TIPO DE MTTO': 'ORDENES_CULMINADAS'})
    
    df_en_ejecucion = df_plan_filtrado[mask_en_ejecucion]
    monthly_en_ejecucion = df_en_ejecucion.groupby(['AÑO', 'MES', 'MES_NOMBRE']).agg({'TIPO DE MTTO': 'count'}).reset_index()
    monthly_en_ejecucion = monthly_en_ejecucion.rename(columns={'TIPO DE MTTO': 'ORDENES_EN_EJECUCION'})
    
    df_retrasadas = df_plan_filtrado[mask_retrasadas]
    monthly_retrasadas = df_retrasadas.groupby(['AÑO', 'MES', 'MES_NOMBRE']).agg({'TIPO DE MTTO': 'count'}).reset_index()
    monthly_retrasadas = monthly_retrasadas.rename(columns={'TIPO DE MTTO': 'ORDENES_RETRASADAS'})
    
    df_proyectadas = df_plan_filtrado[mask_proyectadas]
    monthly_proyectadas = df_proyectadas.groupby(['AÑO', 'MES', 'MES_NOMBRE']).agg({'TIPO DE MTTO': 'count'}).reset_index()
    monthly_proyectadas = monthly_proyectadas.rename(columns={'TIPO DE MTTO': 'ORDENES_PROYECTADAS'})
    
    for _, row in monthly_real_data.iterrows():
        mes = row['MES']
        mask = monthly_data['MES'] == mes
        monthly_data.loc[mask, 'TOTAL_PLANIFICADAS'] = row['TOTAL_PLANIFICADAS']
    
    for _, row in monthly_culminadas.iterrows():
        mes = row['MES']
        mask = monthly_data['MES'] == mes
        monthly_data.loc[mask, 'ORDENES_CULMINADAS'] = row['ORDENES_CULMINADAS']
    
    if not monthly_en_ejecucion.empty:
        for _, row in monthly_en_ejecucion.iterrows():
            mes = row['MES']
            mask = monthly_data['MES'] == mes
            monthly_data.loc[mask, 'ORDENES_EN_EJECUCION'] = row['ORDENES_EN_EJECUCION']
    
    if not monthly_retrasadas.empty:
        for _, row in monthly_retrasadas.iterrows():
            mes = row['MES']
            mask = monthly_data['MES'] == mes
            monthly_data.loc[mask, 'ORDENES_RETRASADAS'] = row['ORDENES_RETRASADAS']
    
    if not monthly_proyectadas.empty:
        for _, row in monthly_proyectadas.iterrows():
            mes = row['MES']
            mask = monthly_data['MES'] == mes
            monthly_data.loc[mask, 'ORDENES_PROYECTADAS'] = row['ORDENES_PROYECTADAS']
    
    monthly_data['CUMPLIMIENTO_PCT'] = monthly_data.apply(
        lambda row: (row['ORDENES_CULMINADAS'] / row['TOTAL_PLANIFICADAS']) * 100 
        if row['TOTAL_PLANIFICADAS'] > 0 else 0, axis=1
    )
    
    monthly_data = monthly_data.sort_values('MES_ORDEN')
    return monthly_data

def get_total_planificadas_mes_actual(df, year=2026):
    if df.empty or 'FECHA_DE_INICIO' not in df.columns or 'TIPO DE MTTO' not in df.columns:
        return 0
    
    mes_actual = datetime.now().month
    año_actual = datetime.now().year
    
    tipos_planificados = ['PREVENTIVO', 'BASADO EN CONDICIÓN', 'MEJORA DE SISTEMA']
    df_plan = df[df['TIPO DE MTTO'].isin(tipos_planificados)].copy()
    
    df_plan_mes_actual = df_plan[
        (df_plan['FECHA_DE_INICIO'].dt.year == año_actual) &
        (df_plan['FECHA_DE_INICIO'].dt.month == mes_actual)
    ]
    
    return len(df_plan_mes_actual)

def get_ordenes_mes_actual(df):
    if df.empty:
        return pd.DataFrame()
    
    mes_actual = datetime.now().month
    año_actual = datetime.now().year
    
    tipos_planificados = ['PREVENTIVO', 'BASADO EN CONDICIÓN', 'MEJORA DE SISTEMA']
    df_plan = df[df['TIPO DE MTTO'].isin(tipos_planificados)].copy()
    
    df_mes_actual = df_plan[
        (df_plan['FECHA_DE_INICIO'].dt.year == año_actual) &
        (df_plan['FECHA_DE_INICIO'].dt.month == mes_actual)
    ]
    
    if df_mes_actual.empty:
        return pd.DataFrame()
    
    if 'STATUS' not in df_mes_actual.columns:
        df_mes_actual['STATUS_NORM'] = 'CULMINADO'
    else:
        df_mes_actual['STATUS_NORM'] = df_mes_actual['STATUS'].astype(str).str.upper().str.strip()
        df_mes_actual.loc[df_mes_actual['STATUS_NORM'].str.contains('CULMINAD'), 'STATUS_NORM'] = 'CULMINADO'
        df_mes_actual.loc[df_mes_actual['STATUS_NORM'].str.contains('PROCESO') | 
                         df_mes_actual['STATUS_NORM'].str.contains('PROGRESO') |
                         df_mes_actual['STATUS_NORM'].str.contains('EJECUCI'), 'STATUS_NORM'] = 'EN PROCESO'
        df_mes_actual.loc[df_mes_actual['STATUS_NORM'].str.contains('PENDIENTE'), 'STATUS_NORM'] = 'PENDIENTE'
    
    fecha_actual = datetime.now().date()
    df_mes_actual['FECHA_INICIO_DATE'] = df_mes_actual['FECHA_DE_INICIO'].dt.date
    df_mes_actual['FECHA_FIN_DATE'] = df_mes_actual['FECHA_DE_FIN'].dt.date
    
    mask_retrasadas = (df_mes_actual['STATUS_NORM'] == 'PENDIENTE') & \
                      (df_mes_actual['FECHA_INICIO_DATE'] < fecha_actual) & \
                      (df_mes_actual['FECHA_FIN_DATE'] < fecha_actual)
    
    mask_en_ejecucion = df_mes_actual['STATUS_NORM'] == 'EN PROCESO'
    
    mask_por_ejecutar = (df_mes_actual['STATUS_NORM'] == 'PENDIENTE') & \
                        ((df_mes_actual['FECHA_INICIO_DATE'] >= fecha_actual) | \
                         (df_mes_actual['FECHA_FIN_DATE'] >= fecha_actual))
    
    mask_ejecutadas = df_mes_actual['STATUS_NORM'] == 'CULMINADO'
    
    df_mes_actual['CATEGORIA'] = 'OTROS'
    df_mes_actual.loc[mask_retrasadas, 'CATEGORIA'] = 'RETRASADA'
    df_mes_actual.loc[mask_en_ejecucion, 'CATEGORIA'] = 'EN EJECUCIÓN'
    df_mes_actual.loc[mask_por_ejecutar, 'CATEGORIA'] = 'POR EJECUTAR EN ENERO'
    df_mes_actual.loc[mask_ejecutadas, 'CATEGORIA'] = 'EJECUTADA'
    
    orden_categorias = ['RETRASADA', 'EN EJECUCIÓN', 'POR EJECUTAR EN ENERO', 'EJECUTADA']
    df_mes_actual['CATEGORIA_ORDEN'] = pd.Categorical(df_mes_actual['CATEGORIA'], categories=orden_categorias, ordered=True)
    df_mes_actual = df_mes_actual.sort_values(['CATEGORIA_ORDEN', 'FECHA_DE_INICIO'])
    
    df_resultado = pd.DataFrame()
    
    columna_ot = None
    posibles_nombres = ['OT', 'N° DE OT', 'N° DE ORDEN', 'NUMERO DE ORDEN', 'N° OT', 'ORDEN']
    
    for nombre in posibles_nombres:
        if nombre in df_mes_actual.columns:
            columna_ot = nombre
            break
    
    if columna_ot:
        try:
            df_mes_actual[columna_ot] = pd.to_numeric(df_mes_actual[columna_ot], errors='coerce')
            df_mes_actual[columna_ot] = df_mes_actual[columna_ot].astype('Int64')
            df_resultado['OT'] = df_mes_actual[columna_ot]
        except Exception as e:
            df_resultado['OT'] = df_mes_actual[columna_ot]
    else:
        df_resultado['OT'] = ''
    
    if 'TIPO DE MTTO' in df_mes_actual.columns:
        df_resultado['TIPO DE MTTO'] = df_mes_actual['TIPO DE MTTO']
    else:
        df_resultado['TIPO DE MTTO'] = ''
    
    if 'EQUIPO' in df_mes_actual.columns:
        df_resultado['EQUIPO'] = df_mes_actual['EQUIPO']
    else:
        df_resultado['EQUIPO'] = ''
    
    if 'FECHA_DE_INICIO' in df_mes_actual.columns:
        df_resultado['FECHA DE INICIO'] = df_mes_actual['FECHA_DE_INICIO'].dt.strftime('%d/%m/%Y')
    else:
        df_resultado['FECHA DE INICIO'] = ''
    
    if 'FECHA_DE_FIN' in df_mes_actual.columns:
        df_resultado['FECHA DE FIN'] = df_mes_actual['FECHA_DE_FIN'].dt.strftime('%d/%m/%Y')
    else:
        df_resultado['FECHA DE FIN'] = ''
    
    if 'STATUS' in df_mes_actual.columns:
        df_resultado['ESTADO'] = df_mes_actual['STATUS']
    else:
        df_resultado['ESTADO'] = df_mes_actual['STATUS_NORM']
    
    df_resultado['CATEGORIA'] = df_mes_actual['CATEGORIA']
    
    return df_resultado

def apply_filters(df, equipo_filter, conjunto_filter, ubicacion_filter, tipo_mtto_filter, fecha_inicio, fecha_fin):
    filtered_df = df.copy()
    
    if equipo_filter != "Todos":
        filtered_df = filtered_df[filtered_df['EQUIPO'].astype(str) == equipo_filter]
    
    if conjunto_filter != "Todos":
        filtered_df = filtered_df[filtered_df['CONJUNTO'].astype(str) == conjunto_filter]
    
    if ubicacion_filter != "Todos":
        if 'UBICACIÓN TÉCNICA' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['UBICACIÓN TÉCNICA'].astype(str) == ubicacion_filter]
    
    if tipo_mtto_filter != "Todos":
        filtered_df = filtered_df[filtered_df['TIPO DE MTTO'].astype(str) == tipo_mtto_filter]
    
    if fecha_inicio is not None and fecha_fin is not None:
        filtered_df = filtered_df[
            (filtered_df['FECHA_DE_INICIO'].dt.date >= fecha_inicio) &
            (filtered_df['FECHA_DE_INICIO'].dt.date <= fecha_fin)
        ]
    
    return filtered_df

def apply_overtime_filters(overtime_df, fecha_inicio, fecha_fin):
    if overtime_df.empty:
        return overtime_df
    
    filtered_overtime = overtime_df.copy()
    
    if fecha_inicio is not None and fecha_fin is not None:
        inicio_datetime = datetime.combine(fecha_inicio, datetime.min.time())
        fin_datetime = datetime.combine(fecha_fin, datetime.max.time())
        
        if 'INICIO_HORAS_EXTRAS' in filtered_overtime.columns:
            if not pd.api.types.is_datetime64_any_dtype(filtered_overtime['INICIO_HORAS_EXTRAS']):
                filtered_overtime['INICIO_HORAS_EXTRAS'] = pd.to_datetime(
                    filtered_overtime['INICIO_HORAS_EXTRAS'], errors='coerce'
                )
            
            filtered_overtime = filtered_overtime[
                (filtered_overtime['INICIO_HORAS_EXTRAS'] >= inicio_datetime) &
                (filtered_overtime['INICIO_HORAS_EXTRAS'] <= fin_datetime)
            ]
    
    return filtered_overtime

def get_current_datetime_spanish():
    now = datetime.now()
    months = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    day = now.day
    month = months[now.month - 1]
    year = now.year
    time_str = now.strftime("%H:%M:%S")
    return f"{day} de {month} de {year}, {time_str}"

def format_date_dd_mm_aaaa(date):
    if isinstance(date, (datetime, pd.Timestamp)):
        return date.strftime('%d/%m/%Y')
    elif isinstance(date, str):
        try:
            return pd.to_datetime(date).strftime('%d/%m/%Y')
        except:
            return date
    else:
        return str(date)

# ============================================
# FUNCIÓN MEJORADA: Generar reporte de horas extras en formato Fortidex
# Replica el "Reporte de Justificación de Sobretiempo" de la imagen
# ============================================

def generate_overtime_report_excel(overtime_data, fecha_inicio, fecha_fin):
    """
    Genera un archivo Excel con:
    - Hoja "HE Semanal por Técnico": tabla semanal por cada técnico con columnas
      SEMANA, DÍA, CANT. HORAS 50%, CANT. HORAS 100%, NÚMERO DE ÓRDENES
    - Hoja "Resumen por Técnico": totales acumulados por técnico
    - Hoja "Detalle Registros": datos crudos completos
    """
    if overtime_data.empty:
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # ── Paleta de colores Fortidex ─────────────────────────────────────
        FORTIDEX_ORANGE = "FF8C00"
        FORTIDEX_BLUE   = "1F4E79"
        HEADER_BLUE     = "BDD7EE"
        ROW_GRAY        = "F2F2F2"
        TOTAL_GREEN     = "E2EFDA"
        SUM_BLUE        = "D6E4F0"
        WHITE           = "FFFFFF"
        BLACK           = "000000"
        TECNICO_HEADER  = "D9E1F2"  # azul muy claro para bloque de técnico

        thin  = Side(style="thin",   color=BLACK)
        thick = Side(style="medium", color="1F4E79")
        thin_border  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
        thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

        def cs(ws, row, col, value="", bold=False, size=10, color=BLACK,
               bg=None, h_align="center", v_align="center",
               wrap=False, border=None, number_format=None):
            """Helper: aplica estilo a una celda."""
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(name="Arial", bold=bold, size=size, color=color)
            c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            if border:
                c.border = border
            if number_format:
                c.number_format = number_format
            return c

        def ms(ws, r1, c1, r2, c2, **kwargs):
            """Helper: merge + estilo."""
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            return cs(ws, r1, c1, **kwargs)

        # ── Preprocesar datos ──────────────────────────────────────────────
        df = overtime_data.copy()

        for col_check, default in [
            ('RESPONSABLE', "SIN NOMBRE"),
            ('INICIO_HORAS_EXTRAS', pd.NaT),
            ('FIN_HORAS_EXTRAS', pd.NaT),
            ('HORAS_EXTRAS', 0.0),
            ('SALDO_HORAS_EXTRAS', 0.0),
        ]:
            if col_check not in df.columns:
                df[col_check] = default

        for col_dt in ['INICIO_HORAS_EXTRAS', 'FIN_HORAS_EXTRAS']:
            if not pd.api.types.is_datetime64_any_dtype(df[col_dt]):
                df[col_dt] = pd.to_datetime(df[col_dt], errors='coerce')

        df['RESPONSABLE'] = df['RESPONSABLE'].astype(str).str.strip()
        df = df[(df['RESPONSABLE'].str.len() > 0) & (df['RESPONSABLE'] != 'nan')]

        def fmt_ot(v):
            if pd.isna(v) or str(v).strip() in ('', 'nan', 'None'):
                return ''
            s = str(v).strip()
            if s.endswith('.0'):
                s = s[:-2]
            try:
                f = float(s)
                return str(int(f)) if f.is_integer() else s
            except:
                return s

        df['OT_FMT'] = df['OT'].apply(fmt_ot) if 'OT' in df.columns else ''

        df_valid = df[df['INICIO_HORAS_EXTRAS'].notna()].copy()
        df_valid['SEMANA_ISO'] = df_valid['INICIO_HORAS_EXTRAS'].dt.isocalendar().week.astype(int)
        df_valid['AÑO_ISO']   = df_valid['INICIO_HORAS_EXTRAS'].dt.isocalendar().year.astype(int)
        df_valid['SEMANA_STR'] = df_valid.apply(
            lambda x: f"{x['AÑO_ISO']}-S{x['SEMANA_ISO']:02d}", axis=1
        )
        df_valid['DOW'] = df_valid['INICIO_HORAS_EXTRAS'].dt.dayofweek  # 0=Lun…6=Dom

        # Días de la semana en orden
        DIAS_ORDEN   = [0, 1, 2, 3, 4, 5, 6]
        DIAS_NOMBRES = {0: 'LUNES', 1: 'MARTES', 2: 'MIÉRCOLES',
                        3: 'JUEVES', 4: 'VIERNES', 5: 'SÁBADO', 6: 'DOMINGO'}
        # Sábado y domingo → 100 %; resto → 50 %
        DIAS_100 = {5, 6}

        df_valid = df_valid.sort_values(['RESPONSABLE', 'AÑO_ISO', 'SEMANA_ISO', 'DOW'])

        # ══════════════════════════════════════════════════════════════════
        # HOJA 1: "HE Semanal por Técnico"
        # Columnas: SEMANA | DÍA | CANT. HORAS 50% | CANT. HORAS 100% | N° OT
        # Solo se muestran los días que caen dentro del rango de fechas filtrado
        # La columna N° OT muestra los códigos de OT del día (no la cantidad)
        # ══════════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet("HE Semanal por Técnico")

        # Anchos de columna — N° OT necesita más espacio para los códigos
        col_w = {1: 20, 2: 14, 3: 22, 4: 22, 5: 40}
        for ci, w in col_w.items():
            ws1.column_dimensions[get_column_letter(ci)].width = w

        # Título general
        ws1.row_dimensions[1].height = 24
        ms(ws1, 1, 1, 1, 5,
           value="REPORTE DE HORAS EXTRAS SEMANAL POR TÉCNICO",
           bold=True, size=13, color=WHITE, bg=FORTIDEX_BLUE,
           h_align="center", border=thick_border)

        # Sub-título período
        ws1.row_dimensions[2].height = 15
        ms(ws1, 2, 1, 2, 5,
           value=f"Período: {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}",
           bold=False, size=10, color=FORTIDEX_BLUE, bg=HEADER_BLUE,
           h_align="center", border=thin_border)

        current_row = 3

        # Convertir fecha_inicio y fecha_fin a date para comparación
        fi_date = fecha_inicio.date() if hasattr(fecha_inicio, 'date') else fecha_inicio
        ff_date = fecha_fin.date()    if hasattr(fecha_fin,    'date') else fecha_fin

        tecnicos_lista = sorted(df_valid['RESPONSABLE'].unique())

        for tecnico in tecnicos_lista:
            df_tec = df_valid[df_valid['RESPONSABLE'] == tecnico].copy()

            # ── Cabecera del bloque del técnico ───────────────────────────
            ws1.row_dimensions[current_row].height = 18
            ms(ws1, current_row, 1, current_row, 5,
               value=f"TÉCNICO:  {tecnico.upper()}",
               bold=True, size=11, color=FORTIDEX_BLUE, bg=TECNICO_HEADER,
               h_align="left", border=thick_border)
            current_row += 1

            # ── Cabecera de columnas ───────────────────────────────────────
            headers = ["SEMANA", "DÍA", "CANT. HORAS 50%", "CANT. HORAS 100%", "N° OT"]
            ws1.row_dimensions[current_row].height = 20
            for ci, hdr in enumerate(headers, start=1):
                cs(ws1, current_row, ci, hdr,
                   bold=True, size=10, color=WHITE, bg=FORTIDEX_ORANGE,
                   border=thin_border)
            current_row += 1

            semanas_tec = sorted(df_tec['SEMANA_STR'].unique())

            total_h50_tec  = 0.0
            total_h100_tec = 0.0

            for semana_str in semanas_tec:
                df_sem = df_tec[df_tec['SEMANA_STR'] == semana_str].copy()

                total_h50_sem  = 0.0
                total_h100_sem = 0.0

                # Calcular lunes y domingo de la semana ISO
                año_iso    = int(df_sem['AÑO_ISO'].iloc[0])
                semana_iso = int(df_sem['SEMANA_ISO'].iloc[0])
                try:
                    # fromisocalendar es el método correcto para semanas ISO
                    # strptime con %W calcula semanas del año civil (no ISO)
                    # y devuelve fechas incorrectas (ej: S08-2026 → 23/02 en vez de 16/02)
                    lunes_d   = datetime.fromisocalendar(int(año_iso), int(semana_iso), 1).date()
                    domingo_d = lunes_d + timedelta(days=6)
                except Exception:
                    lunes_d   = df_sem['INICIO_HORAS_EXTRAS'].min().date()
                    domingo_d = df_sem['INICIO_HORAS_EXTRAS'].max().date()

                # Recortar el rango de días al período filtrado
                # Solo mostrar días que estén dentro de [fi_date, ff_date]
                dia_inicio_semana = max(lunes_d,   fi_date)
                dia_fin_semana    = min(domingo_d, ff_date)

                # Etiqueta de semana con las fechas reales recortadas
                semana_label = (
                    f"S{semana_iso:02d}  "
                    f"{dia_inicio_semana.strftime('%d/%m')} - "
                    f"{dia_fin_semana.strftime('%d/%m/%Y')}"
                )

                # Generar la lista de días dentro del rango para esta semana
                dias_en_rango = []
                for dow in DIAS_ORDEN:
                    fecha_dia = lunes_d + timedelta(days=dow)
                    if fi_date <= fecha_dia <= ff_date:
                        dias_en_rango.append((dow, fecha_dia))

                primera_fila_semana = True

                for i_dia, (dow, fecha_dia) in enumerate(dias_en_rango):
                    dia_nombre = DIAS_NOMBRES[dow]
                    df_dia     = df_sem[df_sem['DOW'] == dow]

                    h_horas = float(df_dia['HORAS_EXTRAS'].sum()) if not df_dia.empty else 0.0
                    bg_row  = WHITE if i_dia % 2 == 0 else ROW_GRAY

                    # Obtener los códigos de OT del día (no la cantidad)
                    if not df_dia.empty and 'OT_FMT' in df_dia.columns:
                        ots = [str(v) for v in df_dia['OT_FMT'].dropna()
                               if str(v).strip() not in ('', 'nan', 'None')]
                        ot_str = " / ".join(ots) if ots else ""
                    else:
                        ot_str = ""

                    ws1.row_dimensions[current_row].height = 14

                    # Columna SEMANA: solo en la primera fila del bloque de semana
                    if primera_fila_semana:
                        cs(ws1, current_row, 1, semana_label, size=9,
                           bg=bg_row, border=thin_border, h_align="center", wrap=True)
                        primera_fila_semana = False
                    else:
                        cs(ws1, current_row, 1, "", size=9, bg=bg_row, border=thin_border)

                    cs(ws1, current_row, 2, dia_nombre, size=9,
                       bg=bg_row, border=thin_border, bold=(dow in DIAS_100))

                    # 50% días hábiles / 100% fin de semana
                    if dow in DIAS_100:
                        h50_val  = 0.0
                        h100_val = h_horas
                    else:
                        h50_val  = h_horas
                        h100_val = 0.0

                    cs(ws1, current_row, 3,
                       round(h50_val, 2) if h50_val > 0 else "",
                       size=9, bg=bg_row, border=thin_border,
                       number_format='0.00' if h50_val > 0 else None)
                    cs(ws1, current_row, 4,
                       round(h100_val, 2) if h100_val > 0 else "",
                       size=9, bg=bg_row, border=thin_border,
                       number_format='0.00' if h100_val > 0 else None)
                    cs(ws1, current_row, 5, ot_str,
                       size=9, bg=bg_row, border=thin_border,
                       h_align="left", wrap=True)

                    total_h50_sem  += h50_val
                    total_h100_sem += h100_val

                    current_row += 1

                # Subtotal de la semana
                ws1.row_dimensions[current_row].height = 14
                cs(ws1, current_row, 1, "SUBTOTAL SEMANA", bold=True, size=9,
                   bg=SUM_BLUE, border=thin_border, h_align="right")
                cs(ws1, current_row, 2, "", size=9, bg=SUM_BLUE, border=thin_border)
                cs(ws1, current_row, 3, round(total_h50_sem, 2) if total_h50_sem > 0 else 0,
                   bold=True, size=9, bg=SUM_BLUE, border=thin_border, number_format='0.00')
                cs(ws1, current_row, 4, round(total_h100_sem, 2) if total_h100_sem > 0 else 0,
                   bold=True, size=9, bg=SUM_BLUE, border=thin_border, number_format='0.00')
                cs(ws1, current_row, 5, "",
                   size=9, bg=SUM_BLUE, border=thin_border)
                current_row += 1

                total_h50_tec  += total_h50_sem
                total_h100_tec += total_h100_sem

            # Total del técnico
            ws1.row_dimensions[current_row].height = 16
            cs(ws1, current_row, 1, f"TOTAL  {tecnico.upper()}", bold=True, size=10,
               bg=TOTAL_GREEN, border=thick_border, h_align="left")
            cs(ws1, current_row, 2, "", bold=True, size=10,
               bg=TOTAL_GREEN, border=thick_border)
            cs(ws1, current_row, 3, round(total_h50_tec, 2),
               bold=True, size=10, bg=TOTAL_GREEN, border=thick_border, number_format='0.00')
            cs(ws1, current_row, 4, round(total_h100_tec, 2),
               bold=True, size=10, bg=TOTAL_GREEN, border=thick_border, number_format='0.00')
            cs(ws1, current_row, 5, "",
               bold=True, size=10, bg=TOTAL_GREEN, border=thick_border)
            current_row += 2  # separador entre técnicos

        # ══════════════════════════════════════════════════════════════════
        # HOJA 2: "Resumen por Técnico"
        # ══════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Resumen por Técnico")

        col_w2 = {1: 35, 2: 20, 3: 20, 4: 18, 5: 18}
        for ci, w in col_w2.items():
            ws2.column_dimensions[get_column_letter(ci)].width = w

        ws2.row_dimensions[1].height = 22
        ms(ws2, 1, 1, 1, 5,
           value="RESUMEN DE HORAS EXTRAS POR TÉCNICO",
           bold=True, size=13, color=WHITE, bg=FORTIDEX_BLUE,
           h_align="center", border=thick_border)

        ws2.row_dimensions[2].height = 15
        ms(ws2, 2, 1, 2, 5,
           value=f"Período: {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}",
           bold=False, size=10, color=FORTIDEX_BLUE, bg=HEADER_BLUE,
           h_align="center", border=thin_border)

        res_headers = ["TÉCNICO", "HORAS 50%", "HORAS 100%", "TOTAL HORAS", "VALOR TOTAL ($)"]
        ws2.row_dimensions[3].height = 18
        for ci, hdr in enumerate(res_headers, start=1):
            cs(ws2, 3, ci, hdr, bold=True, size=10, color=WHITE,
               bg=FORTIDEX_ORANGE, border=thin_border)

        # Calcular resumen agrupado
        def calc_resumen_tecnico(df_t):
            rows = []
            for tec, grp in df_t.groupby('RESPONSABLE'):
                h50  = float(grp[grp['DOW'].isin([0,1,2,3,4])]['HORAS_EXTRAS'].sum())
                h100 = float(grp[grp['DOW'].isin([5,6])]['HORAS_EXTRAS'].sum())
                htot = h50 + h100
                val  = float(grp['SALDO_HORAS_EXTRAS'].sum())
                rows.append({'TECNICO': tec, 'H50': h50, 'H100': h100, 'HTOT': htot, 'VALOR': val})
            return pd.DataFrame(rows).sort_values('HTOT', ascending=False)

        resumen_df = calc_resumen_tecnico(df_valid)
        gt_h50 = gt_h100 = gt_htot = gt_val = 0.0

        for i_r, (_, row_r) in enumerate(resumen_df.iterrows()):
            ri  = i_r + 4
            bgr = WHITE if i_r % 2 == 0 else ROW_GRAY
            ws2.row_dimensions[ri].height = 14

            cs(ws2, ri, 1, str(row_r['TECNICO']), size=10, bg=bgr,
               h_align="left", border=thin_border)
            cs(ws2, ri, 2, round(row_r['H50'],  2), size=10, bg=bgr,
               border=thin_border, number_format='0.00')
            cs(ws2, ri, 3, round(row_r['H100'], 2), size=10, bg=bgr,
               border=thin_border, number_format='0.00')
            cs(ws2, ri, 4, round(row_r['HTOT'], 2), size=10, bg=bgr,
               border=thin_border, number_format='0.00')
            cs(ws2, ri, 5, round(row_r['VALOR'], 2), size=10, bg=bgr,
               border=thin_border, number_format='"$"#,##0.00')

            gt_h50  += row_r['H50']
            gt_h100 += row_r['H100']
            gt_htot += row_r['HTOT']
            gt_val  += row_r['VALOR']

        ri_tot = len(resumen_df) + 4
        ws2.row_dimensions[ri_tot].height = 16
        cs(ws2, ri_tot, 1, "TOTAL GENERAL", bold=True, size=10,
           bg=TOTAL_GREEN, h_align="left", border=thick_border)
        cs(ws2, ri_tot, 2, round(gt_h50,  2), bold=True, size=10,
           bg=TOTAL_GREEN, border=thick_border, number_format='0.00')
        cs(ws2, ri_tot, 3, round(gt_h100, 2), bold=True, size=10,
           bg=TOTAL_GREEN, border=thick_border, number_format='0.00')
        cs(ws2, ri_tot, 4, round(gt_htot, 2), bold=True, size=10,
           bg=TOTAL_GREEN, border=thick_border, number_format='0.00')
        cs(ws2, ri_tot, 5, round(gt_val,  2), bold=True, size=10,
           bg=TOTAL_GREEN, border=thick_border, number_format='"$"#,##0.00')

        # ══════════════════════════════════════════════════════════════════
        # HOJA 3: "Detalle Registros"
        # ══════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Detalle Registros")

        raw_cols = ['RESPONSABLE', 'OT_FMT', 'SEMANA_STR', 'INICIO_HORAS_EXTRAS',
                    'FIN_HORAS_EXTRAS', 'HORAS_EXTRAS', 'SALDO_HORAS_EXTRAS']
        raw_cols = [c for c in raw_cols if c in df_valid.columns]

        raw_names = {
            'RESPONSABLE': 'Técnico', 'OT_FMT': 'N° OT',
            'SEMANA_STR': 'Semana',
            'INICIO_HORAS_EXTRAS': 'Fecha/Hora Inicio',
            'FIN_HORAS_EXTRAS': 'Fecha/Hora Fin',
            'HORAS_EXTRAS': 'Horas Extras', 'SALDO_HORAS_EXTRAS': 'Valor ($)'
        }
        raw_widths = [35, 12, 12, 22, 22, 15, 15]

        for ci, (rc, rw) in enumerate(zip(raw_cols, raw_widths), start=1):
            ws3.column_dimensions[get_column_letter(ci)].width = rw
            cs(ws3, 1, ci, raw_names.get(rc, rc),
               bold=True, size=10, color=WHITE, bg=FORTIDEX_BLUE, border=thin_border)

        for ri, (_, row_raw) in enumerate(df_valid[raw_cols].iterrows(), start=2):
            bgr = WHITE if ri % 2 == 0 else ROW_GRAY
            for ci, rc in enumerate(raw_cols, start=1):
                val = row_raw[rc]
                if rc in ('INICIO_HORAS_EXTRAS', 'FIN_HORAS_EXTRAS'):
                    val = val.strftime('%d/%m/%Y %H:%M') if pd.notna(val) else ''
                elif rc == 'HORAS_EXTRAS':
                    val = round(float(val), 2) if pd.notna(val) else 0
                elif rc == 'SALDO_HORAS_EXTRAS':
                    val = round(float(val), 2) if pd.notna(val) else 0
                cs(ws3, ri, ci, val, size=9, bg=bgr, border=thin_border,
                   h_align="left" if ci == 1 else "center")

        # Guardar y retornar bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    except Exception as e:
        st.error(f"Error al generar reporte Excel: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None


# ============================================
# INTERFAZ PRINCIPAL
# ============================================

def main():
    st.markdown(
    "<h1 style='text-align: center;'>📊 Dashboard de Indicadores de Mantenimiento Mecánico Fortidex</h1>",
    unsafe_allow_html=True
    )
    
    mostrar_info_usuario()
    
    if 'data' not in st.session_state:
        st.session_state.data = pd.DataFrame()
    
    if 'personal_data' not in st.session_state:
        st.session_state.personal_data = pd.DataFrame()
    
    if 'overtime_data' not in st.session_state:
        st.session_state.overtime_data = pd.DataFrame()
    
    if 'last_update' not in st.session_state:
        st.session_state.last_update = None
    
    status_placeholder = st.empty()
    
    if st.session_state.data.empty:
        with status_placeholder.container():
            with st.spinner("Cargando datos desde Google Sheets..."):
                df = load_data_from_google_sheets()

                if not df.empty:
                    st.session_state.data = df
                    st.session_state.last_update = get_current_datetime_spanish()
                    st.success("✅ Datos cargados correctamente desde Google Sheets")
                else:
                    st.error("❌ No se pudieron cargar los datos desde Google Sheets")

        time.sleep(2)
        status_placeholder.empty()
    
    status_personal = st.empty()
    if st.session_state.personal_data.empty:
        with status_personal.container():
            with st.spinner("Cargando datos del personal..."):
                personal_df = load_personal_data_from_google_sheets()

                if not personal_df.empty:
                    st.session_state.personal_data = personal_df
                    st.success("✅ Datos del personal cargados correctamente")
                else:
                    st.warning("⚠️ No se pudieron cargar los datos del personal")

        time.sleep(1)
        status_personal.empty()
    
    status_overtime = st.empty()
    if st.session_state.overtime_data.empty:
        with status_overtime.container():
            with st.spinner("Cargando datos de horas extras..."):
                overtime_df = load_overtime_data_from_google_sheets()

                if not overtime_df.empty:
                    st.session_state.overtime_data = overtime_df
                    st.success("✅ Datos de horas extras cargados correctamente")
                else:
                    st.warning("⚠️ No se pudieron cargar los datos de horas extras")

        time.sleep(1)
        status_overtime.empty()
    
    st.sidebar.title("Opciones")
    
    if not st.session_state.data.empty and st.session_state.last_update:
        st.sidebar.markdown(f"**📅Última actualización:**")
        st.sidebar.markdown(f"`{st.session_state.last_update}`")
        st.sidebar.write(f"**Registros totales:** {len(st.session_state.data)}")
    
    st.sidebar.subheader("Filtros")
    
    if not st.session_state.data.empty:
        min_date = st.session_state.data['FECHA_DE_INICIO'].min().date()
        max_date = st.session_state.data['FECHA_DE_INICIO'].max().date()
        
        st.sidebar.write("**Rango de Fechas**")
        col1, col2 = st.sidebar.columns(2)
        with col1:
            fecha_inicio = st.date_input(
                "Fecha Inicio", value=min_date, min_value=min_date,
                max_value=max_date, key="fecha_inicio"
            )
        with col2:
            fecha_fin = st.date_input(
                "Fecha Fin", value=max_date, min_value=min_date,
                max_value=max_date, key="fecha_fin"
            )
        
        tiempo_disponible, num_dias = calcular_tiempo_disponible(fecha_inicio, fecha_fin)
        fecha_inicio_str = format_date_dd_mm_aaaa(fecha_inicio)
        fecha_fin_str = format_date_dd_mm_aaaa(fecha_fin)
        
        st.sidebar.write(f"**Período seleccionado:**")
        st.sidebar.write(f"**Desde:** {fecha_inicio_str}")
        st.sidebar.write(f"**Hasta:** {fecha_fin_str}")
        st.sidebar.write(f"**Días:** {num_dias}")
        st.sidebar.write(f"**Tiempo disponible:** {tiempo_disponible:,.0f} minutos")
        
        with st.sidebar.expander("📊 Ver cálculo de tiempo disponible"):
            st.write(f"**Fórmula:** Número de días × 24 horas × 60 minutos")
            st.write(f"**Cálculo:** {num_dias} días × 24 horas × 60 minutos = {tiempo_disponible:,.0f} minutos")
            st.write(f"**Ejemplo:** Para 27 días (2026-01-01 a 2026-01-27): 27 × 24 × 60 = 38,880 minutos")
        
        if 'UBICACIÓN TÉCNICA' in st.session_state.data.columns:
            ubicaciones_unique = st.session_state.data['UBICACIÓN TÉCNICA'].dropna().unique().tolist()
            ubicaciones_str = [str(x) for x in ubicaciones_unique]
            ubicaciones = ["Todos"] + sorted(ubicaciones_str)
        else:
            ubicaciones = ["Todos"]
        
        ubicacion_filter = st.sidebar.selectbox("Ubicación Técnica", ubicaciones)
        
        equipos_unique = st.session_state.data['EQUIPO'].unique().tolist()
        equipos_str = [str(x) for x in equipos_unique]
        equipos = ["Todos"] + sorted(equipos_str)
        equipo_filter = st.sidebar.selectbox("Equipo", equipos)
        
        conjuntos_unique = st.session_state.data['CONJUNTO'].unique().tolist()
        conjuntos_str = [str(x) for x in conjuntos_unique]
        conjuntos = ["Todos"] + sorted(conjuntos_str)
        conjunto_filter = st.sidebar.selectbox("Conjunto", conjuntos)
        
        if 'TIPO DE MTTO' in st.session_state.data.columns:
            tipos_mtto_unique = st.session_state.data['TIPO DE MTTO'].dropna().unique().tolist()
            tipos_mtto_str = [str(x) for x in tipos_mtto_unique]
            tipos_mtto = ["Todos"] + sorted(tipos_mtto_str)
        else:
            tipos_mtto = ["Todos"]
        
        tipo_mtto_filter = st.sidebar.selectbox("Tipo de Mtto", tipos_mtto, key="tipo_mtto_filter")
        
        filtered_data = apply_filters(st.session_state.data, equipo_filter, conjunto_filter, 
                                      ubicacion_filter, tipo_mtto_filter, fecha_inicio, fecha_fin)
        
        filtered_overtime = apply_overtime_filters(st.session_state.overtime_data, fecha_inicio, fecha_fin)
        
        filtered_data = clean_responsable_column(filtered_data)
        filtered_overtime = clean_responsable_column(filtered_overtime)
        
        st.sidebar.subheader("Estado")
        st.sidebar.write(f"**Registros filtrados:** {len(filtered_data)}")
        st.sidebar.write(f"**Horas extras registros:** {len(filtered_overtime)}")
        st.sidebar.write(f"**Equipos únicos:** {len(filtered_data['EQUIPO'].unique())}")
        if not filtered_data.empty and 'FECHA_DE_INICIO' in filtered_data.columns:
            min_date_filtered = filtered_data['FECHA_DE_INICIO'].min()
            max_date_filtered = filtered_data['FECHA_DE_INICIO'].max()
            min_date_str = format_date_dd_mm_aaaa(min_date_filtered)
            max_date_str = format_date_dd_mm_aaaa(max_date_filtered)
            st.sidebar.write(f"**Período datos:** {min_date_str} a {max_date_str}")
        
        st.markdown("""
        <style>
        .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
            font-size: 1.2rem;
            font-weight: 600;
        }
        .stTabs [data-baseweb="tab-list"] { gap: 8px; }
        .stTabs [data-baseweb="tab-list"] button { padding: 12px 24px; }
        </style>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10 = st.tabs([
            "Planta", "TFS", "TR", "TFC", "Tipo de Mtto", "Confiabilidad", 
            "Horas Personal Técnico", "Costos Horas Extras Personal Técnico",
            "Cumplimiento del Plan", "Reporte Horas Extras"
        ])
        
        metrics = calculate_metrics(filtered_data, fecha_inicio, fecha_fin, filtered_overtime)
        weekly_data = get_weekly_data(filtered_data, fecha_inicio, fecha_fin)
        reliability_metrics = calculate_reliability_metrics(filtered_data, fecha_inicio, fecha_fin)
        weekly_emergency_data = get_weekly_emergency_data(filtered_data)
        weekly_tech_data = get_weekly_technician_hours(filtered_data, filtered_overtime)
        accumulated_tech_data = get_accumulated_technician_hours(filtered_data, filtered_overtime)
        weekly_costs, accumulated_costs, mensaje_calculo = calculate_overtime_costs_from_details(filtered_overtime, st.session_state.personal_data)
        monthly_plan_data = get_monthly_plan_data(st.session_state.data, year=2026)
        total_planificadas_mes_actual = get_total_planificadas_mes_actual(st.session_state.data, year=2026)
        ordenes_mes_actual = get_ordenes_mes_actual(st.session_state.data)
        
        # ── TAB 1: Planta ──────────────────────────────────────────────────
        with tab1:
            st.header("🏭 Dashboard de Planta - Vista Consolidada")
            if not filtered_data.empty:
                st.subheader("📊 Indicadores Clave de Desempeño")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    disponibilidad = metrics.get('disponibilidad', 0)
                    fig_dispo = go.Figure(go.Indicator(
                        mode="gauge+number+delta", value=disponibilidad,
                        domain={'x': [0, 1], 'y': [0, 1]},
                        title={'text': "Disponibilidad", 'font': {'size': 18}},
                        delta={'reference': 80, 'increasing': {'color': "green"}},
                        gauge={'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
                               'bar': {'color': "darkblue"}, 'bgcolor': "white",
                               'borderwidth': 2, 'bordercolor': "gray",
                               'steps': [{'range': [0, 60], 'color': '#FF0000'},
                                         {'range': [60, 80], 'color': '#FFD700'},
                                         {'range': [80, 100], 'color': '#32CD32'}],
                               'threshold': {'line': {'color': "black", 'width': 4},
                                             'thickness': 0.75, 'value': disponibilidad}}
                    ))
                    fig_dispo.update_layout(height=300)
                    st.plotly_chart(fig_dispo, use_container_width=True)
                    if disponibilidad >= 80:
                        st.success("✅ Excelente")
                    elif disponibilidad >= 60:
                        st.warning("⚠️ Regular")
                    else:
                        st.error("❌ Crítico")
                
                with col2:
                    if not monthly_plan_data.empty and 'TOTAL_PLANIFICADAS' in monthly_plan_data.columns:
                        total_planificadas = monthly_plan_data['TOTAL_PLANIFICADAS'].sum()
                        total_culminadas = monthly_plan_data['ORDENES_CULMINADAS'].sum()
                        cumplimiento = (total_culminadas / total_planificadas * 100) if total_planificadas > 0 else 0
                    else:
                        cumplimiento = 0
                    
                    fig_cumplimiento = go.Figure(go.Indicator(
                        mode="gauge+number+delta", value=cumplimiento,
                        domain={'x': [0, 1], 'y': [0, 1]},
                        title={'text': "Cumplimiento Plan", 'font': {'size': 18}},
                        delta={'reference': 80, 'increasing': {'color': "green"}},
                        gauge={'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
                               'bar': {'color': "darkblue"}, 'bgcolor': "white",
                               'borderwidth': 2, 'bordercolor': "gray",
                               'steps': [{'range': [0, 70], 'color': '#FF0000'},
                                         {'range': [70, 90], 'color': '#FFD700'},
                                         {'range': [90, 100], 'color': '#32CD32'}],
                               'threshold': {'line': {'color': "black", 'width': 4},
                                             'thickness': 0.75, 'value': cumplimiento}}
                    ))
                    fig_cumplimiento.update_layout(height=300)
                    st.plotly_chart(fig_cumplimiento, use_container_width=True)
                    if cumplimiento >= 90:
                        st.success("✅ Excelente")
                    elif cumplimiento >= 70:
                        st.warning("⚠️ Regular")
                    else:
                        st.error("❌ Crítico")
                
                with col3:
                    mtbf = reliability_metrics.get('mtbf_emergency', 0) if reliability_metrics else 0
                    fig_mtbf = go.Figure(go.Indicator(
                        mode="gauge+number", value=mtbf,
                        domain={'x': [0, 1], 'y': [0, 1]},
                        title={'text': "MTBF (min)", 'font': {'size': 18}},
                        gauge={'axis': {'range': [0, 10000], 'tickwidth': 1, 'tickcolor': "darkblue"},
                               'bar': {'color': "darkblue"}, 'bgcolor': "white",
                               'borderwidth': 2, 'bordercolor': "gray",
                               'steps': [{'range': [0, 3000], 'color': '#FF0000'},
                                         {'range': [3001, 6000], 'color': '#FFD700'},
                                         {'range': [6001, 10000], 'color': '#32CD32'}],
                               'threshold': {'line': {'color': "black", 'width': 4},
                                             'thickness': 0.75, 'value': mtbf}}
                    ))
                    fig_mtbf.update_layout(height=300)
                    st.plotly_chart(fig_mtbf, use_container_width=True)
                    if mtbf >= 6001:
                        st.success("✅ Excelente")
                    elif mtbf >= 3001:
                        st.warning("⚠️ Regular")
                    else:
                        st.error("❌ Crítico")
                
                with col4:
                    mttr = reliability_metrics.get('mttr_emergency', 0) if reliability_metrics else 0
                    mttr_normalizado = min(mttr, 500)
                    fig_mttr = go.Figure(go.Indicator(
                        mode="gauge+number", value=mttr_normalizado,
                        domain={'x': [0, 1], 'y': [0, 1]},
                        title={'text': "MTTR (min)", 'font': {'size': 18}},
                        gauge={'axis': {'range': [0, 500], 'tickwidth': 1, 'tickcolor': "darkblue"},
                               'bar': {'color': "darkblue"}, 'bgcolor': "white",
                               'borderwidth': 2, 'bordercolor': "gray",
                               'steps': [{'range': [0, 120], 'color': '#32CD32'},
                                         {'range': [120, 240], 'color': '#FFD700'},
                                         {'range': [240, 500], 'color': '#FF0000'}],
                               'threshold': {'line': {'color': "black", 'width': 4},
                                             'thickness': 0.75, 'value': mttr_normalizado}}
                    ))
                    fig_mttr.update_layout(height=300)
                    st.plotly_chart(fig_mttr, use_container_width=True)
                    if mttr <= 120:
                        st.success("✅ Excelente")
                    elif mttr <= 240:
                        st.warning("⚠️ Regular")
                    else:
                        st.error("❌ Crítico")
                
                st.subheader("📈 Métricas Operativas")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    td = metrics.get('td', 0)
                    st.metric("Tiempo Disponible", f"{td:,.0f}", "minutos",
                             help=f"Calculado para {num_dias} días entre {fecha_inicio_str} y {fecha_fin_str}")
                with col2:
                    st.metric("Tiempo Operativo", f"{metrics.get('to', 0):,.0f}", "minutos")
                with col3:
                    st.metric("TFS Total", f"{metrics.get('tfs', 0):,.0f}", "minutos")
                with col4:
                    horas_extras = metrics.get('horas_extras_acumuladas', 0)
                    st.metric("Horas Extras Acum", f"{horas_extras/60:,.1f}", "horas")
                
                st.subheader("📊 Evolución de Indicadores")
                col1, col2 = st.columns(2)
                
                with col1:
                    if not weekly_data.empty and 'DISPO_SEMANAL' in weekly_data.columns:
                        try:
                            fig = px.line(weekly_data, x='SEMANA_STR', y='DISPO_SEMANAL', 
                                        title='📈 Disponibilidad por Semana',
                                        labels={'SEMANA_STR': 'Semana', 'DISPO_SEMANAL': 'Disponibilidad (%)'},
                                        color_discrete_sequence=['#32CD32'])
                            fig.update_traces(mode='lines+markers', line_width=3)
                            fig.add_hrect(y0=80, y1=100, line_width=0, fillcolor="green", opacity=0.1)
                            fig.add_hrect(y0=60, y1=80, line_width=0, fillcolor="yellow", opacity=0.1)
                            fig.add_hrect(y0=0, y1=60, line_width=0, fillcolor="red", opacity=0.1)
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as e:
                            st.error(f"Error al crear gráfico de disponibilidad: {str(e)[:100]}")
                    else:
                        st.info("No hay datos de disponibilidad semanal")
                
                with col2:
                    if not weekly_data.empty:
                        try:
                            fig = go.Figure()
                            if 'TFS_MIN' in weekly_data.columns:
                                fig.add_trace(go.Bar(x=weekly_data['SEMANA_STR'], y=weekly_data['TFS_MIN'],
                                                    name='TFS', marker_color='#FF6B6B'))
                            if 'TR_MIN' in weekly_data.columns:
                                fig.add_trace(go.Scatter(x=weekly_data['SEMANA_STR'], y=weekly_data['TR_MIN'],
                                                        name='TR', mode='lines+markers',
                                                        line=dict(color='#FFD700', width=3)))
                            fig.update_layout(title='📊 TFS y TR por Semana',
                                            xaxis_title='Semana', yaxis_title='Minutos',
                                            hovermode='x unified')
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as e:
                            st.error(f"Error en gráfico TFS/TR: {str(e)[:100]}")
                    else:
                        st.info("No hay datos de TFS y TR semanal")
                
                col1, col2 = st.columns(2)
                with col1:
                    if not weekly_emergency_data.empty and 'NUM_ORDENES_EMERGENCIA' in weekly_emergency_data.columns:
                        fig = px.bar(weekly_emergency_data, x='SEMANA_STR', y='NUM_ORDENES_EMERGENCIA',
                                    title='🚨 Correctivos de Emergencia por Semana',
                                    labels={'SEMANA_STR': 'Semana', 'NUM_ORDENES_EMERGENCIA': 'N° de Órdenes'},
                                    color='NUM_ORDENES_EMERGENCIA', color_continuous_scale='Reds')
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos de correctivos de emergencia")
                
                with col2:
                    if not weekly_data.empty and 'TFC_MIN' in weekly_data.columns:
                        fig = go.Figure()
                        fig.add_trace(go.Scatter(x=weekly_data['SEMANA_STR'], y=weekly_data['TFC_MIN'],
                                                name='TFC', mode='lines+markers',
                                                line=dict(color='#4ECDC4', width=3)))
                        fig.update_layout(title='🔧 TFC por Semana',
                                        xaxis_title='Semana', yaxis_title='TFC (min)')
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos de TFC semanal")
                
                st.subheader("🔍 Análisis por Equipo y Conjunto")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    filtered_afecta = filtered_data[filtered_data['PRODUCCION_AFECTADA'] == 'SI']
                    if not filtered_afecta.empty and 'EQUIPO' in filtered_afecta.columns and 'TFS_MIN' in filtered_afecta.columns:
                        tfs_por_equipo = filtered_afecta.groupby('EQUIPO')['TFS_MIN'].sum().reset_index()
                        tfs_por_equipo = tfs_por_equipo.nlargest(5, 'TFS_MIN')
                        if not tfs_por_equipo.empty:
                            fig = px.bar(tfs_por_equipo, x='EQUIPO', y='TFS_MIN',
                                        title='🛠️ TFS Top 5 Equipos',
                                        labels={'EQUIPO': 'Equipo', 'TFS_MIN': 'TFS (min)'},
                                        color='TFS_MIN', color_continuous_scale='Reds')
                            fig.update_layout(showlegend=False)
                            st.plotly_chart(fig, use_container_width=True)
                
                with col2:
                    if not filtered_afecta.empty and 'CONJUNTO' in filtered_afecta.columns and 'TR_MIN' in filtered_afecta.columns:
                        tr_por_conjunto = filtered_afecta.groupby('CONJUNTO')['TR_MIN'].sum().reset_index()
                        tr_por_conjunto = tr_por_conjunto.nlargest(5, 'TR_MIN')
                        if not tr_por_conjunto.empty:
                            fig = px.bar(tr_por_conjunto, x='CONJUNTO', y='TR_MIN',
                                        title='🔧 TR Top 5 Conjuntos',
                                        labels={'CONJUNTO': 'Conjunto', 'TR_MIN': 'TR (min)'},
                                        color='TR_MIN', color_continuous_scale='Oranges')
                            fig.update_layout(showlegend=False)
                            st.plotly_chart(fig, use_container_width=True)
                
                with col3:
                    if 'TIPO DE MTTO' in filtered_data.columns:
                        tipo_mtto_counts = filtered_data['TIPO DE MTTO'].value_counts().reset_index()
                        tipo_mtto_counts.columns = ['TIPO_MTTO', 'COUNT']
                        if not tipo_mtto_counts.empty:
                            fig = px.pie(tipo_mtto_counts, values='COUNT', names='TIPO_MTTO',
                                        title='📋 Distribución Tipo Mtto', hole=0.4,
                                        color_discrete_sequence=px.colors.qualitative.Set3)
                            fig.update_traces(textposition='inside', textinfo='percent+label')
                            st.plotly_chart(fig, use_container_width=True)
                
                with col4:
                    if not accumulated_costs.empty and 'COSTO_TOTAL' in accumulated_costs.columns:
                        top_tecnicos = accumulated_costs.nlargest(5, 'COSTO_TOTAL')
                        if not top_tecnicos.empty:
                            fig = px.bar(top_tecnicos, x='TECNICO', y='COSTO_TOTAL',
                                        title='💰 Costos Horas Extras Top 5',
                                        labels={'TECNICO': 'Técnico', 'COSTO_TOTAL': 'Costo Total ($)'},
                                        color='COSTO_TOTAL', color_continuous_scale='Greens')
                            fig.update_layout(xaxis_tickangle=-45, showlegend=False)
                            st.plotly_chart(fig, use_container_width=True)
                
                st.subheader("📅 Cumplimiento del Plan 2026")
                
                if not monthly_plan_data.empty and 'ORDENES_CULMINADAS' in monthly_plan_data.columns:
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        fig = go.Figure()
                        fig.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'],
                                            y=monthly_plan_data['ORDENES_RETRASADAS'],
                                            name='Retrasadas', marker_color='#FFA500'))
                        fig.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'],
                                            y=monthly_plan_data['ORDENES_EN_EJECUCION'],
                                            name='En Ejecución', marker_color='#FFD700'))
                        fig.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'],
                                            y=monthly_plan_data['ORDENES_CULMINADAS'],
                                            name='Culminadas', marker_color='#32CD32'))
                        fig.update_layout(barmode='stack', title='📊 Estado de Órdenes por Mes',
                                        xaxis_title='Mes', yaxis_title='Número de Órdenes', height=400)
                        st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        if 'CUMPLIMIENTO_PCT' in monthly_plan_data.columns:
                            fig = px.line(monthly_plan_data, x='MES_NOMBRE', y='CUMPLIMIENTO_PCT',
                                        title='📈 Porcentaje de Cumplimiento',
                                        labels={'MES_NOMBRE': 'Mes', 'CUMPLIMIENTO_PCT': 'Cumplimiento (%)'},
                                        markers=True)
                            fig.update_traces(line_color='#32CD32', line_width=3)
                            fig.add_hrect(y0=90, y1=100, line_width=0, fillcolor="green", opacity=0.1)
                            fig.add_hrect(y0=70, y1=90, line_width=0, fillcolor="yellow", opacity=0.1)
                            fig.add_hrect(y0=0, y1=70, line_width=0, fillcolor="red", opacity=0.1)
                            fig.update_layout(height=400)
                            st.plotly_chart(fig, use_container_width=True)
                
                st.subheader("📋 Resumen Estadístico")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    total_fallas = reliability_metrics.get('total_fallas_emergency', 0) if reliability_metrics else 0
                    st.metric("Total Fallas", f"{total_fallas}",
                            help="Número de correctivos de emergencia")
                with col2:
                    tr_horas = accumulated_tech_data['TR_HORAS'].sum() if not accumulated_tech_data.empty and 'TR_HORAS' in accumulated_tech_data.columns else 0
                    st.metric("Horas Normales Técnicos", f"{tr_horas:,.1f}", "horas")
                with col3:
                    costo_total = accumulated_costs['COSTO_TOTAL'].sum() if not accumulated_costs.empty and 'COSTO_TOTAL' in accumulated_costs.columns else 0
                    st.metric("Costo Total Horas Extras", f"${costo_total:,.2f}")
                with col4:
                    td = metrics.get('td', 0)
                    to = metrics.get('to', 0)
                    eficiencia_global = (to / td * 100) if td > 0 else 0
                    st.metric("Eficiencia Global", f"{eficiencia_global:.1f}%")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 2: TFS ────────────────────────────────────────────────────
        with tab2:
            st.header("Análisis de TFS")
            if not filtered_data.empty:
                filtered_afecta = filtered_data[filtered_data['PRODUCCION_AFECTADA'] == 'SI']
                col1, col2 = st.columns(2)
                with col1:
                    if not weekly_data.empty:
                        fig = px.line(weekly_data, x='SEMANA_STR', y='TFS_MIN',
                                     title='TFS por Semana (Minutos)',
                                     labels={'SEMANA_STR': 'Semana', 'TFS_MIN': 'TFS (min)'})
                        fig.update_traces(line_color=COLOR_PALETTE['pastel'][1], mode='lines+markers')
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos semanales para mostrar")
                with col2:
                    tfs_por_equipo = filtered_afecta.groupby('EQUIPO')['TFS_MIN'].sum().reset_index()
                    tfs_por_equipo = tfs_por_equipo.sort_values('TFS_MIN', ascending=False).head(10)
                    if not tfs_por_equipo.empty:
                        fig = px.bar(tfs_por_equipo, x='EQUIPO', y='TFS_MIN',
                                    title='TFS por Equipo',
                                    labels={'EQUIPO': 'Equipo', 'TFS_MIN': 'TFS (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][1])
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos de TFS por equipo")
                
                tfs_por_conjunto = filtered_afecta.groupby('CONJUNTO')['TFS_MIN'].sum().reset_index()
                tfs_por_conjunto = tfs_por_conjunto.sort_values('TFS_MIN', ascending=False).head(10)
                if not tfs_por_conjunto.empty:
                    fig = px.bar(tfs_por_conjunto, x='CONJUNTO', y='TFS_MIN',
                                title='TFS por Conjunto',
                                labels={'CONJUNTO': 'Conjunto', 'TFS_MIN': 'TFS (min)'})
                    fig.update_traces(marker_color=COLOR_PALETTE['pastel'][1])
                    st.plotly_chart(fig, use_container_width=True)
                
                if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                    tfs_por_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA')['TFS_MIN'].sum().reset_index()
                    tfs_por_ubicacion = tfs_por_ubicacion.sort_values('TFS_MIN', ascending=False).head(10)
                    if not tfs_por_ubicacion.empty:
                        fig = px.bar(tfs_por_ubicacion, x='UBICACIÓN TÉCNICA', y='TFS_MIN',
                                    title='TFS por Ubicación Técnica',
                                    labels={'UBICACIÓN TÉCNICA': 'Ubicación Técnica', 'TFS_MIN': 'TFS (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][1])
                        st.plotly_chart(fig, use_container_width=True)
                
                st.subheader("Resúmenes TFS")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write("**Resumen TFS por Equipo**")
                    resumen_equipo = filtered_afecta.groupby('EQUIPO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_equipo, use_container_width=True)
                with col2:
                    st.write("**Resumen TFS por Conjunto**")
                    resumen_conjunto = filtered_afecta.groupby('CONJUNTO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_conjunto.head(10), use_container_width=True)
                with col3:
                    st.write("**Resumen TFS por Ubicación Técnica**")
                    if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                        resumen_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                        st.dataframe(resumen_ubicacion.head(10), use_container_width=True)
                    else:
                        st.info("No hay datos de ubicación técnica")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 3: TR ─────────────────────────────────────────────────────
        with tab3:
            st.header("Análisis de TR")
            if not filtered_data.empty:
                filtered_afecta = filtered_data[filtered_data['PRODUCCION_AFECTADA'] == 'SI']
                col1, col2 = st.columns(2)
                with col1:
                    if not weekly_data.empty:
                        fig = px.line(weekly_data, x='SEMANA_STR', y='TR_MIN',
                                     title='TR por Semana (Minutos)',
                                     labels={'SEMANA_STR': 'Semana', 'TR_MIN': 'TR (min)'})
                        fig.update_traces(line_color=COLOR_PALETTE['pastel'][2], mode='lines+markers')
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos semanales para mostrar")
                with col2:
                    tr_por_equipo = filtered_afecta.groupby('EQUIPO')['TR_MIN'].sum().reset_index()
                    tr_por_equipo = tr_por_equipo.sort_values('TR_MIN', ascending=False).head(10)
                    if not tr_por_equipo.empty:
                        fig = px.bar(tr_por_equipo, x='EQUIPO', y='TR_MIN',
                                    title='TR por Equipo',
                                    labels={'EQUIPO': 'Equipo', 'TR_MIN': 'TR (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][2])
                        st.plotly_chart(fig, use_container_width=True)
                
                tr_por_conjunto = filtered_afecta.groupby('CONJUNTO')['TR_MIN'].sum().reset_index()
                tr_por_conjunto = tr_por_conjunto.sort_values('TR_MIN', ascending=False).head(15)
                if not tr_por_conjunto.empty:
                    fig = px.bar(tr_por_conjunto, x='CONJUNTO', y='TR_MIN',
                                title='Pareto TR por Conjunto',
                                labels={'CONJUNTO': 'Conjunto', 'TR_MIN': 'TR (min)'})
                    fig.update_traces(marker_color=COLOR_PALETTE['pastel'][2])
                    st.plotly_chart(fig, use_container_width=True)
                
                if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                    tr_por_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA')['TR_MIN'].sum().reset_index()
                    tr_por_ubicacion = tr_por_ubicacion.sort_values('TR_MIN', ascending=False).head(10)
                    if not tr_por_ubicacion.empty:
                        fig = px.bar(tr_por_ubicacion, x='UBICACIÓN TÉCNICA', y='TR_MIN',
                                    title='TR por Ubicación Técnica',
                                    labels={'UBICACIÓN TÉCNICA': 'Ubicación Técnica', 'TR_MIN': 'TR (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][2])
                        st.plotly_chart(fig, use_container_width=True)
                
                st.subheader("Resúmenes TR")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write("**Resumen TR por Equipo**")
                    resumen_equipo = filtered_afecta.groupby('EQUIPO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_equipo, use_container_width=True)
                with col2:
                    st.write("**Resumen TR por Conjunto**")
                    resumen_conjunto = filtered_afecta.groupby('CONJUNTO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_conjunto.head(10), use_container_width=True)
                with col3:
                    st.write("**Resumen TR por Ubicación Técnica**")
                    if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                        resumen_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                        st.dataframe(resumen_ubicacion.head(10), use_container_width=True)
                    else:
                        st.info("No hay datos de ubicación técnica")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 4: TFC ────────────────────────────────────────────────────
        with tab4:
            st.header("Análisis de TFC")
            if not filtered_data.empty:
                filtered_afecta = filtered_data[filtered_data['PRODUCCION_AFECTADA'] == 'SI']
                col1, col2 = st.columns(2)
                with col1:
                    if not weekly_data.empty:
                        fig = px.line(weekly_data, x='SEMANA_STR', y='TFC_MIN',
                                     title='TFC por Semana (Minutos)',
                                     labels={'SEMANA_STR': 'Semana', 'TFC_MIN': 'TFC (min)'})
                        fig.update_traces(line_color=COLOR_PALETTE['pastel'][3], mode='lines+markers')
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos semanales para mostrar")
                with col2:
                    tfc_por_equipo = filtered_afecta.groupby('EQUIPO')['TFC_MIN'].sum().reset_index()
                    tfc_por_equipo = tfc_por_equipo.sort_values('TFC_MIN', ascending=False).head(10)
                    if not tfc_por_equipo.empty:
                        fig = px.bar(tfc_por_equipo, x='EQUIPO', y='TFC_MIN',
                                    title='TFC por Equipo',
                                    labels={'EQUIPO': 'Equipo', 'TFC_MIN': 'TFC (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][3])
                        st.plotly_chart(fig, use_container_width=True)
                
                tfc_por_conjunto = filtered_afecta.groupby('CONJUNTO')['TFC_MIN'].sum().reset_index()
                tfc_por_conjunto = tfc_por_conjunto.sort_values('TFC_MIN', ascending=False).head(15)
                if not tfc_por_conjunto.empty:
                    fig = px.bar(tfc_por_conjunto, x='CONJUNTO', y='TFC_MIN',
                                title='Pareto TFC por Conjunto',
                                labels={'CONJUNTO': 'Conjunto', 'TFC_MIN': 'TFC (min)'})
                    fig.update_traces(marker_color=COLOR_PALETTE['pastel'][3])
                    st.plotly_chart(fig, use_container_width=True)
                
                if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                    tfc_por_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA')['TFC_MIN'].sum().reset_index()
                    tfc_por_ubicacion = tfc_por_ubicacion.sort_values('TFC_MIN', ascending=False).head(10)
                    if not tfc_por_ubicacion.empty:
                        fig = px.bar(tfc_por_ubicacion, x='UBICACIÓN TÉCNICA', y='TFC_MIN',
                                    title='TFC por Ubicación Técnica',
                                    labels={'UBICACIÓN TÉCNICA': 'Ubicación Técnica', 'TFC_MIN': 'TFC (min)'})
                        fig.update_traces(marker_color=COLOR_PALETTE['pastel'][3])
                        st.plotly_chart(fig, use_container_width=True)
                
                st.subheader("Resúmenes TFC")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write("**Resumen TFC por Equipo**")
                    resumen_equipo = filtered_afecta.groupby('EQUIPO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_equipo, use_container_width=True)
                with col2:
                    st.write("**Resumen TFC por Conjunto**")
                    resumen_conjunto = filtered_afecta.groupby('CONJUNTO').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                    st.dataframe(resumen_conjunto.head(10), use_container_width=True)
                with col3:
                    st.write("**Resumen TFC por Ubicación Técnica**")
                    if 'UBICACIÓN TÉCNICA' in filtered_afecta.columns:
                        resumen_ubicacion = filtered_afecta.groupby('UBICACIÓN TÉCNICA').agg({'TFS_MIN': 'sum', 'TR_MIN': 'sum', 'TFC_MIN': 'sum'}).reset_index()
                        st.dataframe(resumen_ubicacion.head(10), use_container_width=True)
                    else:
                        st.info("No hay datos de ubicación técnica")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 5: Tipo de Mantenimiento ──────────────────────────────────
        with tab5:
            st.header("Análisis por Tipo de Mantenimiento")
            if not filtered_data.empty:
                if 'STATUS' in filtered_data.columns:
                    filtered_data_mtto = filtered_data.copy()
                    filtered_data_mtto['STATUS_NORM'] = filtered_data_mtto['STATUS'].astype(str).str.upper().str.strip()
                    filtered_data_mtto.loc[filtered_data_mtto['STATUS_NORM'].str.contains('CULMINAD'), 'STATUS_NORM'] = 'CULMINADO'
                    filtered_data_mtto = filtered_data_mtto[filtered_data_mtto['STATUS_NORM'] == 'CULMINADO']
                else:
                    st.warning("⚠️ No se encontró la columna 'STATUS'. Mostrando todos los datos sin filtrar por estado.")
                    filtered_data_mtto = filtered_data
                
                metrics_mtto = calculate_metrics(filtered_data_mtto, fecha_inicio, fecha_fin, filtered_overtime)
                
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Mantenimiento Preventivo", f"{metrics_mtto.get('mp_pct', 0):.1f}%")
                with col2:
                    st.metric("Mant. Basado en Condición", f"{metrics_mtto.get('mbc_pct', 0):.1f}%")
                with col3:
                    st.metric("Correctivo Programado", f"{metrics_mtto.get('mcp_pct', 0):.1f}%")
                with col4:
                    st.metric("Correctivo de Emergencia", f"{metrics_mtto.get('mce_pct', 0):.1f}%")
                with col5:
                    st.metric("Mejora de Sistema", f"{metrics_mtto.get('mms_pct', 0):.1f}%")
                
                col1, col2 = st.columns(2)
                with col1:
                    if 'FECHA_DE_INICIO' in filtered_data_mtto.columns and 'TIPO DE MTTO' in filtered_data_mtto.columns and 'TR_MIN' in filtered_data_mtto.columns:
                        df_weekly_mtto = filtered_data_mtto.copy()
                        df_weekly_mtto['SEMANA'] = df_weekly_mtto['FECHA_DE_INICIO'].dt.isocalendar().week
                        df_weekly_mtto['AÑO'] = df_weekly_mtto['FECHA_DE_INICIO'].dt.year
                        df_weekly_mtto['SEMANA_STR'] = df_weekly_mtto.apply(
                            lambda x: f"{x['AÑO']}-S{x['SEMANA']:02d}", axis=1
                        )
                        try:
                            tipo_mtto_semana = df_weekly_mtto.groupby(['SEMANA_STR', 'TIPO DE MTTO'])['TR_MIN'].sum().reset_index()
                            if not tipo_mtto_semana.empty:
                                tipo_mtto_semana['AÑO_NUM'] = tipo_mtto_semana['SEMANA_STR'].str.extract(r'(\d{4})').astype(int)
                                tipo_mtto_semana['SEMANA_NUM'] = tipo_mtto_semana['SEMANA_STR'].str.extract(r'-S(\d{1,2})').astype(int)
                                tipo_mtto_semana = tipo_mtto_semana.sort_values(['AÑO_NUM', 'SEMANA_NUM'])
                                semanas_ordenadas = tipo_mtto_semana['SEMANA_STR'].unique().tolist()
                                tipos_mtto_unicos = tipo_mtto_semana['TIPO DE MTTO'].unique()
                                tipos_ordenados = []
                                for tipo in ['PREVENTIVO', 'BASADO EN CONDICIÓN', 'CORRECTIVO PROGRAMADO', 'CORRECTIVO DE EMERGENCIA', 'MEJORA DE SISTEMA']:
                                    if tipo in tipos_mtto_unicos:
                                        tipos_ordenados.append(tipo)
                                for tipo in tipos_mtto_unicos:
                                    if tipo not in tipos_ordenados:
                                        tipos_ordenados.append(tipo)
                                
                                fig = px.bar(tipo_mtto_semana, x='SEMANA_STR', y='TR_MIN', color='TIPO DE MTTO',
                                            title='Tipo de Mantenimiento por Semana (Barras Apiladas)',
                                            labels={'SEMANA_STR': 'Semana', 'TR_MIN': 'Tiempo (min)'},
                                            color_discrete_map=COLOR_PALETTE['tipo_mtto'],
                                            category_orders={'SEMANA_STR': semanas_ordenadas, 'TIPO DE MTTO': tipos_ordenados})
                                if len(semanas_ordenadas) > 8:
                                    fig.update_xaxes(tickangle=45)
                                st.plotly_chart(fig, use_container_width=True)
                        except Exception as e:
                            st.error(f"Error al crear gráfico de barras: {str(e)[:100]}")
                
                with col2:
                    if 'TIPO DE MTTO' in filtered_data_mtto.columns and 'TR_MIN' in filtered_data_mtto.columns:
                        try:
                            tipo_mtto_totals = filtered_data_mtto.groupby('TIPO DE MTTO')['TR_MIN'].sum().reset_index()
                            if not tipo_mtto_totals.empty:
                                total_minutos = tipo_mtto_totals['TR_MIN'].sum()
                                tipo_mtto_totals['PORCENTAJE'] = (tipo_mtto_totals['TR_MIN'] / total_minutos * 100).round(1)
                                tipo_mtto_totals['HORAS'] = (tipo_mtto_totals['TR_MIN'] / 60).round(1)
                                tipo_mtto_totals['TEXTO_PORCENTAJE'] = tipo_mtto_totals['PORCENTAJE'].apply(lambda x: f"{x:.1f}%")
                                
                                color_map_extendido = COLOR_PALETTE['tipo_mtto'].copy()
                                colores_adicionales = ['#FFA500', '#800080', '#008000', '#FF69B4', '#00CED1']
                                tipos_existentes = tipo_mtto_totals['TIPO DE MTTO'].unique()
                                tipos_ordenados_filtrados = [t for t in ['PREVENTIVO', 'BASADO EN CONDICIÓN', 'CORRECTIVO PROGRAMADO', 'CORRECTIVO DE EMERGENCIA', 'MEJORA DE SISTEMA'] if t in tipos_existentes]
                                for tipo in tipos_existentes:
                                    if tipo not in tipos_ordenados_filtrados:
                                        tipos_ordenados_filtrados.append(tipo)
                                for i, tipo in enumerate(tipos_ordenados_filtrados):
                                    if tipo not in color_map_extendido:
                                        color_map_extendido[tipo] = colores_adicionales[i % len(colores_adicionales)]
                                
                                fig = px.treemap(tipo_mtto_totals,
                                                path=['TIPO DE MTTO'], values='TR_MIN',
                                                title='Distribución del Mantenimiento (Órdenes CULMINADAS)',
                                                color='TIPO DE MTTO', color_discrete_map=color_map_extendido,
                                                hover_data={'TR_MIN': True, 'HORAS': True, 'PORCENTAJE': True},
                                                custom_data=['TIPO DE MTTO', 'TR_MIN', 'HORAS', 'PORCENTAJE'])
                                fig.update_traces(
                                    textinfo="label+text",
                                    text=tipo_mtto_totals['TEXTO_PORCENTAJE'],
                                    texttemplate="<b>%{label}</b><br>%{text}",
                                    textposition="middle center",
                                    textfont=dict(size=16, color='white'),
                                    hovertemplate="<b>%{customdata[0]}</b><br>Minutos: %{value:.0f}<br>Horas: %{customdata[2]:.1f}<br>Porcentaje: %{customdata[3]:.1f}%<extra></extra>",
                                    marker=dict(cornerradius=5)
                                )
                                fig.update_layout(margin=dict(t=50, l=25, r=25, b=25), height=500)
                                st.plotly_chart(fig, use_container_width=True)
                        except Exception as e:
                            st.error(f"Error al crear gráfico de treemap: {str(e)[:200]}")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 6: Confiabilidad ──────────────────────────────────────────
        with tab6:
            st.header("Indicadores de Confiabilidad")
            if not filtered_data.empty:
                if reliability_metrics:
                    col1, col2, col3, col4, col5, col6 = st.columns(6)
                    with col1:
                        st.metric("Total Fallas", f"{reliability_metrics.get('total_fallas_emergency', 0):,.0f}")
                    with col2:
                        st.metric("Total Fallas con parada", f"{reliability_metrics.get('total_fallas_emergency_con_parada', 0):,.0f}")
                    with col3:
                        st.metric("MTBF", f"{reliability_metrics.get('mtbf_emergency', 0):,.1f}", "minutos")
                    with col4:
                        st.metric("MTTF", f"{reliability_metrics.get('mttf_emergency', 0):,.1f}", "minutos")
                    with col5:
                        st.metric("MTTR", f"{reliability_metrics.get('mttr_emergency', 0):,.1f}", "minutos")
                    with col6:
                        mantenibilidad_pct = reliability_metrics.get('mantenibilidad_pct', 0)
                        st.metric("Mantenibilidad", f"{mantenibilidad_pct:.1f}%")
                else:
                    st.info("No hay datos de correctivos de emergencia para calcular las métricas")
                
                col1, col2 = st.columns(2)
                with col1:
                    if not weekly_emergency_data.empty:
                        fig = px.bar(weekly_emergency_data, x='SEMANA_STR', y='NUM_ORDENES_EMERGENCIA',
                                    title='Total de Fallas por Semana (Correctivos de Emergencia)',
                                    labels={'SEMANA_STR': 'Semana', 'NUM_ORDENES_EMERGENCIA': 'N° de Órdenes de Emergencia'},
                                    color='NUM_ORDENES_EMERGENCIA', color_continuous_scale='Reds')
                        fig.update_layout(showlegend=False)
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos semanales de correctivos de emergencia")
                with col2:
                    if not weekly_emergency_data.empty:
                        fig = px.line(weekly_emergency_data, x='SEMANA_STR', y='MTTR_SEMANAL',
                                     title='MTTR por Semana (Correctivos de Emergencia)',
                                     labels={'SEMANA_STR': 'Semana', 'MTTR_SEMANAL': 'MTTR (min)'},
                                     markers=True)
                        fig.update_traces(line_color='#FFA500', mode='lines+markers', line_width=3)
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.info("No hay datos semanales para calcular MTTR")
                
                st.subheader("Distribución de Correctivos de Emergencia")
                emergency_data = filtered_data[filtered_data['TIPO DE MTTO'] == 'CORRECTIVO DE EMERGENCIA']
                if not emergency_data.empty:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write("**Distribución por Equipo (Top 10)**")
                        emergencia_por_equipo = emergency_data.groupby('EQUIPO').size().reset_index(name='CANTIDAD')
                        emergencia_por_equipo = emergencia_por_equipo.sort_values('CANTIDAD', ascending=False).head(10)
                        emergencia_por_equipo = emergencia_por_equipo.reset_index(drop=True)
                        emergencia_por_equipo.insert(0, 'LUGAR', range(1, len(emergencia_por_equipo) + 1))
                        emergencia_por_equipo['LUGAR'] = emergencia_por_equipo['LUGAR'].astype(str) + '°'
                        emergencia_por_equipo = emergencia_por_equipo.rename(columns={'CANTIDAD': 'CANTIDAD DE FALLA'})
                        emergencia_por_equipo = emergencia_por_equipo[['LUGAR', 'EQUIPO', 'CANTIDAD DE FALLA']]
                        st.dataframe(emergencia_por_equipo, use_container_width=True)
                    with col2:
                        st.write("**Distribución por Conjunto (Top 10)**")
                        emergencia_por_conjunto = emergency_data.groupby('CONJUNTO').size().reset_index(name='CANTIDAD')
                        emergencia_por_conjunto = emergencia_por_conjunto.sort_values('CANTIDAD', ascending=False).head(10)
                        emergencia_por_conjunto = emergencia_por_conjunto.reset_index(drop=True)
                        emergencia_por_conjunto.insert(0, 'LUGAR', range(1, len(emergencia_por_conjunto) + 1))
                        emergencia_por_conjunto['LUGAR'] = emergencia_por_conjunto['LUGAR'].astype(str) + '°'
                        emergencia_por_conjunto = emergencia_por_conjunto.rename(columns={'CANTIDAD': 'CANTIDAD DE FALLA'})
                        emergencia_por_conjunto = emergencia_por_conjunto[['LUGAR', 'CONJUNTO', 'CANTIDAD DE FALLA']]
                        st.dataframe(emergencia_por_conjunto, use_container_width=True)
                else:
                    st.info("No hay registros de correctivos de emergencia en el período seleccionado")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados")
        
        # ── TAB 7: Horas Personal Técnico ─────────────────────────────────
        with tab7:
            st.header("👷 Análisis de Horas del Personal Técnico")
            if not filtered_data.empty:
                if 'RESPONSABLE' not in filtered_data.columns:
                    st.warning("⚠️ La columna 'RESPONSABLE' no está presente en los datos.")
                else:
                    data_with_responsible_separado = separar_tecnicos(filtered_data)
                    if data_with_responsible_separado.empty:
                        st.info("No hay datos con responsable asignado para mostrar.")
                    else:
                        if not weekly_tech_data.empty:
                            tecnicos_unicos = weekly_tech_data['RESPONSABLE'].unique()
                            colores_tecnicos = {}
                            colores_disponibles = COLOR_PALETTE['pastel'] + ['#FFA07A', '#20B2AA', '#778899', '#B0C4DE', '#FFB6C1', '#98FB98', '#DDA0DD', '#FFE4B5']
                            for i, tecnico in enumerate(tecnicos_unicos):
                                colores_tecnicos[tecnico] = colores_disponibles[i % len(colores_disponibles)]
                            
                            st.subheader("📊 Horas Normales por Técnico")
                            col1, col2 = st.columns(2)
                            with col1:
                                semanas_ordenadas = sorted(weekly_tech_data['SEMANA_STR'].unique())
                                fig = px.bar(weekly_tech_data, x='SEMANA_STR', y='TR_HORAS', color='RESPONSABLE',
                                            title='Horas Normales por Semana (por Técnico)',
                                            labels={'SEMANA_STR': 'Semana', 'TR_HORAS': 'Horas Normales', 'RESPONSABLE': 'Técnico'},
                                            color_discrete_map=colores_tecnicos,
                                            category_orders={'SEMANA_STR': semanas_ordenadas})
                                fig.update_layout(barmode='stack')
                                st.plotly_chart(fig, use_container_width=True)
                            with col2:
                                horas_normales_acumuladas = data_with_responsible_separado.groupby('RESPONSABLE')['TR_MIN'].sum().reset_index()
                                horas_normales_acumuladas['TR_HORAS'] = horas_normales_acumuladas['TR_MIN'] / 60
                                horas_normales_acumuladas = horas_normales_acumuladas.sort_values('TR_HORAS', ascending=False)
                                if not horas_normales_acumuladas.empty:
                                    horas_normales_acumuladas['LABEL'] = horas_normales_acumuladas.apply(
                                        lambda x: f"{x['RESPONSABLE']}: {x['TR_HORAS']:.1f} horas", axis=1)
                                    fig = px.pie(horas_normales_acumuladas, values='TR_HORAS', names='LABEL',
                                                title='Distribución de Horas Normales Acumuladas',
                                                color='RESPONSABLE', color_discrete_map=colores_tecnicos)
                                    fig.update_traces(textposition='inside', textinfo='percent+label')
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            st.subheader("⏰ Horas Extras por Técnico (desde DETALLE_HE)")
                            weekly_overtime = get_weekly_overtime_data(filtered_overtime)
                            if not weekly_overtime.empty:
                                col1, col2 = st.columns(2)
                                with col1:
                                    semanas_ordenadas = sorted(weekly_overtime['SEMANA_STR'].unique())
                                    fig = px.bar(weekly_overtime, x='SEMANA_STR', y='H_EXTRA_HORAS', color='RESPONSABLE',
                                                title='Horas Extras por Semana (por Técnico)',
                                                labels={'SEMANA_STR': 'Semana', 'H_EXTRA_HORAS': 'Horas Extras', 'RESPONSABLE': 'Técnico'},
                                                color_discrete_map=colores_tecnicos,
                                                category_orders={'SEMANA_STR': semanas_ordenadas})
                                    fig.update_layout(barmode='stack')
                                    st.plotly_chart(fig, use_container_width=True)
                                with col2:
                                    horas_extras_acumuladas = get_accumulated_overtime_data(filtered_overtime)
                                    if not horas_extras_acumuladas.empty and 'H_EXTRA_HORAS' in horas_extras_acumuladas.columns:
                                        horas_extras_acumuladas = horas_extras_acumuladas[horas_extras_acumuladas['H_EXTRA_HORAS'] > 0]
                                        horas_extras_acumuladas = horas_extras_acumuladas.sort_values('H_EXTRA_HORAS', ascending=False)
                                        if not horas_extras_acumuladas.empty:
                                            horas_extras_acumuladas['LABEL'] = horas_extras_acumuladas.apply(
                                                lambda x: f"{x['RESPONSABLE']}: {x['H_EXTRA_HORAS']:.1f} horas", axis=1)
                                            fig = px.pie(horas_extras_acumuladas, values='H_EXTRA_HORAS', names='LABEL',
                                                        title='Distribución de Horas Extras Acumuladas',
                                                        color='RESPONSABLE', color_discrete_map=colores_tecnicos)
                                            fig.update_traces(textposition='inside', textinfo='percent+label')
                                            st.plotly_chart(fig, use_container_width=True)
                            else:
                                st.info("No hay datos de horas extras por técnico para mostrar.")
                        else:
                            st.info("No hay datos semanales por técnico para mostrar.")
            else:
                st.info("No hay datos para mostrar con los filtros seleccionados.")
        
        # ── TAB 8: Costos Horas Extras ────────────────────────────────────
        with tab8:
            st.header("💰 Costos de Horas Extras del Personal Técnico (desde DETALLE_HE)")
            if not filtered_overtime.empty:
                weekly_costs, accumulated_costs, mensaje_calculo = calculate_overtime_costs_from_details(filtered_overtime, st.session_state.personal_data)
                
                if weekly_costs.empty or accumulated_costs.empty:
                    with st.expander("🔍 Depuración - Ver detalles de los datos", expanded=True):
                        st.subheader("Registros de horas extras encontrados en DETALLE_HE")
                        if not filtered_overtime.empty:
                            st.write(f"**Total de registros de horas extras:** {len(filtered_overtime)}")
                            columnas = ['RESPONSABLE', 'OT', 'INICIO_HORAS_EXTRAS', 'HORAS_EXTRAS', 'SALDO_HORAS_EXTRAS']
                            existing_columns = [col for col in columnas if col in filtered_overtime.columns]
                            st.dataframe(filtered_overtime[existing_columns].head(20), use_container_width=True)
                        else:
                            st.warning("No se encontraron registros en DETALLE_HE para el período seleccionado")
                else:
                    show_detailed_costs_info(weekly_costs, accumulated_costs, st.session_state.personal_data)
                    
                    tecnicos_unicos = list(weekly_costs['TECNICO'].unique())
                    colores_tecnicos = {}
                    colores_disponibles = COLOR_PALETTE['pastel'] + ['#FFA07A', '#20B2AA', '#778899', '#B0C4DE', '#FFB6C1', '#98FB98', '#DDA0DD', '#FFE4B5']
                    for i, tecnico in enumerate(tecnicos_unicos):
                        colores_tecnicos[tecnico] = colores_disponibles[i % len(colores_disponibles)]
                    
                    st.subheader("📈 Evolución de Costos por Semana")
                    col1, col2 = st.columns(2)
                    with col1:
                        semanas_ordenadas = sorted(weekly_costs['SEMANA_STR'].unique())
                        fig = px.bar(weekly_costs, x='SEMANA_STR', y='COSTO_TOTAL', color='TECNICO',
                                    title='Costos de Horas Extras por Semana (USD)',
                                    labels={'SEMANA_STR': 'Semana', 'COSTO_TOTAL': 'Costo Total (USD)', 'TECNICO': 'Técnico'},
                                    color_discrete_map=colores_tecnicos,
                                    category_orders={'SEMANA_STR': semanas_ordenadas})
                        fig.update_layout(barmode='stack')
                        st.plotly_chart(fig, use_container_width=True)
                    with col2:
                        fig = px.bar(weekly_costs, x='SEMANA_STR', y='H_EXTRA_HORAS', color='TECNICO',
                                    title='Horas Extras por Semana',
                                    labels={'SEMANA_STR': 'Semana', 'H_EXTRA_HORAS': 'Horas Extras', 'TECNICO': 'Técnico'},
                                    color_discrete_map=colores_tecnicos,
                                    category_orders={'SEMANA_STR': semanas_ordenadas})
                        fig.update_layout(barmode='stack')
                        st.plotly_chart(fig, use_container_width=True)
                    
                    st.subheader("📊 Análisis de Distribución")
                    col1, col2 = st.columns(2)
                    with col1:
                        pie_data = accumulated_costs.copy()
                        pie_data['PORCENTAJE'] = (pie_data['COSTO_TOTAL'] / pie_data['COSTO_TOTAL'].sum()) * 100
                        pie_data['LABEL'] = pie_data.apply(
                            lambda x: f"{x['TECNICO']}: ${x['COSTO_TOTAL']:,.2f} ({x['PORCENTAJE']:.1f}%)", axis=1)
                        fig = px.pie(pie_data, values='COSTO_TOTAL', names='LABEL',
                                    title='Distribución de Costos de Horas Extras',
                                    color='TECNICO', color_discrete_map=colores_tecnicos)
                        fig.update_traces(textposition='inside', textinfo='percent+label')
                        st.plotly_chart(fig, use_container_width=True)
                    with col2:
                        fig = px.bar(accumulated_costs.sort_values('COSTO_TOTAL', ascending=True),
                                    y='TECNICO', x='COSTO_TOTAL',
                                    title='Costos Acumulados por Técnico',
                                    labels={'TECNICO': 'Técnico', 'COSTO_TOTAL': 'Costo Total (USD)'},
                                    color='TECNICO', color_discrete_map=colores_tecnicos, orientation='h')
                        fig.update_traces(texttemplate='$%{x:,.2f}', textposition='outside')
                        st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No hay datos de horas extras para el período seleccionado.")
        
        # ── TAB 9: Cumplimiento del Plan ──────────────────────────────────
        with tab9:
            st.header("📋 Cumplimiento del Plan de Mantenimiento 2026")
            
            with st.expander("ℹ️ **Información sobre el cálculo del cumplimiento**", expanded=False):
                st.markdown("""
                ### 📊 **Cálculo del Cumplimiento del Plan**
                #### **DEFINICIÓN CORREGIDA DE ÓRDENES RETRASADAS:**
                - Estado: 'PENDIENTE', Fecha de inicio < hoy, Fecha de fin < hoy
                """)
            
            st.subheader("📅 Indicador del Mes Actual")
            mes_actual = datetime.now().month
            año_actual = datetime.now().year
            col_actual1, col_actual2 = st.columns(2)
            with col_actual1:
                st.metric("Total Planificadas del Mes", f"{total_planificadas_mes_actual}")
            with col_actual2:
                nombre_mes = monthly_plan_data[monthly_plan_data['MES'] == mes_actual]['MES_NOMBRE'].iloc[0] if not monthly_plan_data[monthly_plan_data['MES'] == mes_actual].empty else mes_actual
                st.metric("Mes de referencia", f"{nombre_mes} {año_actual}")
            
            if not monthly_plan_data.empty:
                total_planificadas = float(monthly_plan_data['TOTAL_PLANIFICADAS'].sum())
                total_culminadas = float(monthly_plan_data['ORDENES_CULMINADAS'].sum())
                total_en_ejecucion = float(monthly_plan_data['ORDENES_EN_EJECUCION'].sum())
                total_retrasadas = float(monthly_plan_data['ORDENES_RETRASADAS'].sum())
                cumplimiento_general = (total_culminadas / total_planificadas * 100) if total_planificadas > 0 else 0
                
                if cumplimiento_general >= 90:
                    estado_plan = "🟢 Excelente"; estado_color = "green"
                elif cumplimiento_general >= 70:
                    estado_plan = "🟡 Bueno"; estado_color = "orange"
                elif cumplimiento_general >= 50:
                    estado_plan = "🟠 Regular"; estado_color = "#FF8C00"
                else:
                    estado_plan = "🔴 Crítico"; estado_color = "red"
                
                st.subheader("📊 Indicadores Generales del Plan 2026")
                col1, col2, col3, col4, col5, col6 = st.columns(6)
                with col1:
                    st.metric("Total Planificadas hasta ayer", f"{total_planificadas}")
                with col2:
                    st.metric("Órdenes Culminadas", f"{total_culminadas}")
                with col3:
                    st.metric("Órdenes en Ejecución", f"{total_en_ejecucion}")
                with col4:
                    st.metric("Órdenes Retrasadas", f"{total_retrasadas}")
                with col5:
                    st.metric("Cumplimiento", f"{cumplimiento_general:.1f}%")
                with col6:
                    st.markdown(f"**Estado del Plan**")
                    st.markdown(f"<h3 style='color:{estado_color};'>{estado_plan}</h3>", unsafe_allow_html=True)
                
                fig1 = go.Figure()
                fig1.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'], y=monthly_plan_data['ORDENES_RETRASADAS'],
                                     name='Retrasadas', marker_color=COLOR_PALETTE['estado_orden']['RETRASADAS'],
                                     text=monthly_plan_data['ORDENES_RETRASADAS'], textposition='inside',
                                     textfont=dict(size=15, color='black')))
                fig1.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'], y=monthly_plan_data['ORDENES_EN_EJECUCION'],
                                     name='En Ejecución', marker_color=COLOR_PALETTE['estado_orden']['EN EJECUCIÓN'],
                                     text=monthly_plan_data['ORDENES_EN_EJECUCION'], textposition='inside',
                                     textfont=dict(size=15, color='black')))
                fig1.add_trace(go.Bar(x=monthly_plan_data['MES_NOMBRE'], y=monthly_plan_data['ORDENES_CULMINADAS'],
                                     name='Culminadas', marker_color=COLOR_PALETTE['estado_orden']['CULMINADAS'],
                                     text=monthly_plan_data['ORDENES_CULMINADAS'], textposition='inside',
                                     textfont=dict(size=15, color='black')))
                fig1.add_trace(go.Scatter(x=monthly_plan_data['MES_NOMBRE'], y=monthly_plan_data['TOTAL_PLANIFICADAS'],
                                         name='Total Planificado hasta ayer', mode='lines+markers',
                                         line=dict(color=COLOR_PALETTE['estado_orden']['TOTAL_PLANIFICADAS'], width=3, dash='dash')))
                
                for i, row in monthly_plan_data.iterrows():
                    if row['TOTAL_PLANIFICADAS'] > 0:
                        cumplimiento_mensual = row['CUMPLIMIENTO_PCT']
                        color_texto = 'green' if cumplimiento_mensual >= 90 else ('orange' if cumplimiento_mensual >= 80 else ('red' if cumplimiento_mensual < 70 else '#FF8C00'))
                        fig1.add_annotation(x=row['MES_NOMBRE'], y=row['TOTAL_PLANIFICADAS'] + (row['TOTAL_PLANIFICADAS'] * 0.05),
                                           text=f"{cumplimiento_mensual:.0f}%", showarrow=False,
                                           font=dict(size=20, color=color_texto, weight='bold'), yshift=5)
                
                fig1.update_layout(title='Distribución de Órdenes por Mes', xaxis_title='Mes',
                                  yaxis_title='Número de Órdenes', barmode='stack', hovermode='x unified', height=500,
                                  legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                st.plotly_chart(fig1, use_container_width=True)
                
                st.subheader("📋 Tabla de Órdenes del Mes Actual")
                if not ordenes_mes_actual.empty:
                    resumen_categorias = ordenes_mes_actual['CATEGORIA'].value_counts().reset_index()
                    resumen_categorias.columns = ['Categoría', 'Cantidad']
                    orden_categorias = ['RETRASADA', 'EN EJECUCIÓN', 'POR EJECUTAR EN ENERO', 'EJECUTADA']
                    resumen_categorias['Categoría'] = pd.Categorical(resumen_categorias['Categoría'], categories=orden_categorias, ordered=True)
                    resumen_categorias = resumen_categorias.sort_values('Categoría')
                    
                    cols_resumen = st.columns(4)
                    for idx, (_, row) in enumerate(resumen_categorias.iterrows()):
                        with cols_resumen[idx % 4]:
                            if row['Categoría'] == 'RETRASADA':
                                st.metric("Órdenes Retrasadas", f"{row['Cantidad']}", delta_color="inverse")
                            elif row['Categoría'] == 'EN EJECUCIÓN':
                                st.metric("Órdenes en Ejecución", f"{row['Cantidad']}")
                            elif row['Categoría'] == 'POR EJECUTAR EN ENERO':
                                st.metric("Por Ejecutar en Enero", f"{row['Cantidad']}")
                            elif row['Categoría'] == 'EJECUTADA':
                                st.metric("Órdenes Ejecutadas", f"{row['Cantidad']}")
                    
                    columna_ot = None
                    for nombre in ['OT', 'N° DE OT', 'N° DE ORDEN', 'NUMERO DE ORDEN', 'N° OT', 'ORDEN']:
                        if nombre in ordenes_mes_actual.columns:
                            columna_ot = nombre
                            break
                    
                    if columna_ot:
                        columnas_solicitadas = [columna_ot, 'TIPO DE MTTO', 'EQUIPO', 'FECHA DE INICIO', 'FECHA DE FIN', 'ESTADO', 'CATEGORIA']
                        columnas_existentes = [col for col in columnas_solicitadas if col in ordenes_mes_actual.columns]
                        tabla_mostrar = ordenes_mes_actual[columnas_existentes].copy()
                        
                        if columna_ot in tabla_mostrar.columns:
                            try:
                                tabla_mostrar[columna_ot] = pd.to_numeric(tabla_mostrar[columna_ot], errors='coerce')
                                tabla_mostrar[columna_ot] = tabla_mostrar[columna_ot].fillna(0).astype('Int64')
                                tabla_mostrar[columna_ot] = tabla_mostrar[columna_ot].replace(0, '')
                            except Exception as e:
                                st.warning(f"No se pudo convertir la columna {columna_ot} a entero: {e}")
                        
                        if columna_ot in tabla_mostrar.columns:
                            tabla_mostrar = tabla_mostrar.rename(columns={columna_ot: 'N° DE OT'})
                            nuevas_columnas = ['N° DE OT'] + [col for col in tabla_mostrar.columns if col != 'N° DE OT']
                            tabla_mostrar = tabla_mostrar[nuevas_columnas]
                        
                        def color_categoria(val):
                            if val == 'RETRASADA': return 'background-color: #FFA500; color: black; font-weight: bold'
                            elif val == 'EN EJECUCIÓN': return 'background-color: #FFD700; color: black; font-weight: bold'
                            elif val == 'POR EJECUTAR EN ENERO': return 'background-color: #52b3f3; color: black; font-weight: bold'
                            elif val == 'EJECUTADA': return 'background-color: #32CD32; color: black; font-weight: bold'
                            return ''
                        
                        styled_table = tabla_mostrar.style.applymap(color_categoria, subset=['CATEGORIA'])
                        st.dataframe(styled_table, use_container_width=True, height=400, hide_index=True)
                        
                        csv = tabla_mostrar.to_csv(index=False).encode('utf-8')
                        st.download_button(label="📥 Descargar tabla como CSV", data=csv,
                                          file_name=f"ordenes_mes_actual_{datetime.now().strftime('%Y%m%d')}.csv",
                                          mime="text/csv")
                else:
                    st.info("No hay órdenes planificadas para el mes actual.")
                
                st.subheader("📈 Cumplimiento por Mes")
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=monthly_plan_data['MES_NOMBRE'], y=monthly_plan_data['CUMPLIMIENTO_PCT'],
                                         mode='lines+markers+text', name='% Cumplimiento',
                                         line=dict(color='#32CD32', width=3), marker=dict(size=10, color='#32CD32'),
                                         text=[f"{val:.0f}%" for val in monthly_plan_data['CUMPLIMIENTO_PCT']],
                                         textposition='top center', textfont=dict(size=12, color='black')))
                fig2.add_hline(y=80, line_dash="dash", line_color="orange",
                              annotation_text="Objetivo 80%", annotation_position="bottom right")
                fig2.add_hline(y=90, line_dash="dash", line_color="green",
                              annotation_text="Excelente 90%", annotation_position="top right")
                fig2.update_layout(title='Porcentaje de Cumplimiento por Mes',
                                  xaxis_title='Mes', yaxis_title='Cumplimiento (%)',
                                  yaxis_range=[0, 105], height=400, showlegend=True)
                st.plotly_chart(fig2, use_container_width=True)
        
        # ── TAB 10: Reporte Horas Extras ──────────────────────────────────
        with tab10:
            st.header("📊 Reporte de Horas Extras")
            st.markdown("""
            > El reporte se genera en formato **Reporte de Justificación de Sobretiempo** de Fortidex,
            > organizado por semana y empleado, con columnas de **H. ENTRADA**, **H. SALIDA**, **50%**, **100%**
            > y **ACTIVIDAD/JUSTIFICACIÓN**, igual al documento oficial de la empresa.
            """)
            
            st.subheader("🔍 Filtros para el Reporte")
            
            if not st.session_state.overtime_data.empty and 'INICIO_HORAS_EXTRAS' in st.session_state.overtime_data.columns:
                min_date_overtime = st.session_state.overtime_data['INICIO_HORAS_EXTRAS'].min().date()
                max_date_overtime = st.session_state.overtime_data['INICIO_HORAS_EXTRAS'].max().date()
            else:
                min_date_overtime = datetime(2026, 1, 1).date()
                max_date_overtime = datetime(2026, 12, 31).date()
            
            col1, col2 = st.columns(2)
            with col1:
                report_fecha_inicio = st.date_input(
                    "Fecha de inicio para el reporte",
                    value=min_date_overtime, min_value=min_date_overtime,
                    max_value=max_date_overtime, key="report_fecha_inicio"
                )
            with col2:
                report_fecha_fin = st.date_input(
                    "Fecha de fin para el reporte",
                    value=max_date_overtime, min_value=min_date_overtime,
                    max_value=max_date_overtime, key="report_fecha_fin"
                )
            
            filtered_overtime_report = apply_overtime_filters(
                st.session_state.overtime_data, report_fecha_inicio, report_fecha_fin
            )
            
            st.info(f"Filtros aplicados: {len(filtered_overtime_report)} registros encontrados entre {report_fecha_inicio} y {report_fecha_fin}")
            
            if not filtered_overtime_report.empty:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Registros", f"{len(filtered_overtime_report)}")
                with col2:
                    if 'HORAS_EXTRAS' in filtered_overtime_report.columns:
                        total_horas = filtered_overtime_report['HORAS_EXTRAS'].sum()
                        st.metric("Horas Extras Totales", f"{total_horas:.2f} horas")
                    else:
                        st.metric("Horas Extras Totales", "N/A")
                with col3:
                    if 'SALDO_HORAS_EXTRAS' in filtered_overtime_report.columns:
                        total_costo = filtered_overtime_report['SALDO_HORAS_EXTRAS'].sum()
                        st.metric("Costo Total", f"${total_costo:,.2f}")
                    else:
                        st.metric("Costo Total", "N/A")
                with col4:
                    if 'RESPONSABLE' in filtered_overtime_report.columns:
                        tecnicos_unicos = filtered_overtime_report['RESPONSABLE'].nunique()
                        st.metric("Técnicos Involucrados", f"{tecnicos_unicos}")
                    else:
                        st.metric("Técnicos Involucrados", "N/A")
                
                st.subheader("👁️ Vista Previa de los Datos")
                preview_cols = ['RESPONSABLE', 'OT', 'INICIO_HORAS_EXTRAS', 'HORAS_EXTRAS', 'SALDO_HORAS_EXTRAS']
                existing_cols = [col for col in preview_cols if col in filtered_overtime_report.columns]
                
                if existing_cols:
                    preview_df = filtered_overtime_report[existing_cols].copy()
                    if 'OT' in preview_df.columns:
                        def format_ot_as_int(ot_value):
                            if pd.isna(ot_value) or ot_value == '': return ''
                            ot_str = str(ot_value).strip().replace(' ', '')
                            if ot_str.replace('-', '', 1).isdigit(): return ot_str
                            if '.' in ot_str:
                                try:
                                    num = float(ot_str)
                                    return str(int(num)) if num.is_integer() else f"{num:.2f}"
                                except: return ot_str
                            return ot_str
                        preview_df['OT'] = preview_df['OT'].apply(format_ot_as_int)
                    if 'INICIO_HORAS_EXTRAS' in preview_df.columns:
                        preview_df['INICIO_HORAS_EXTRAS'] = preview_df['INICIO_HORAS_EXTRAS'].dt.strftime('%d/%m/%Y %H:%M')
                    if 'HORAS_EXTRAS' in preview_df.columns:
                        preview_df['HORAS_EXTRAS'] = preview_df['HORAS_EXTRAS'].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "0.00")
                    if 'SALDO_HORAS_EXTRAS' in preview_df.columns:
                        preview_df['SALDO_HORAS_EXTRAS'] = preview_df['SALDO_HORAS_EXTRAS'].apply(
                            lambda x: f"${x:,.2f}" if pd.notnull(x) and x != 0 else "$0.00")
                    
                    st.dataframe(preview_df.head(20), use_container_width=True,
                                column_config={
                                    "OT": st.column_config.TextColumn("N° OT", width="small"),
                                    "RESPONSABLE": st.column_config.TextColumn("Técnico", width="medium"),
                                    "INICIO_HORAS_EXTRAS": st.column_config.TextColumn("Fecha/Hora Inicio", width="medium"),
                                    "HORAS_EXTRAS": st.column_config.NumberColumn("Horas Extras", format="%.2f horas"),
                                    "SALDO_HORAS_EXTRAS": st.column_config.TextColumn("Valor ($)", width="small")
                                })
                    if len(preview_df) > 20:
                        st.info(f"Mostrando 20 de {len(preview_df)} registros. El reporte completo se incluirá en el archivo Excel.")
                
                # ── Descarga del Excel ──────────────────────────────────────
                st.subheader("📥 Descargar Reporte en Excel")
                
                st.markdown("""
                ### **El archivo Excel generado incluye:**
                | Hoja | Contenido |
                |------|-----------|
                | **Reporte Sobretiempo** | Formato oficial Fortidex: encabezado institucional, tabla por empleado con DIA, FECHA, H. ENTRADA, H. SALIDA, 50%, 100%, ACTIVIDAD — organizado por semana |
                | **Resumen por Técnico** | Totales de horas y valor en dólares por técnico |
                | **Detalle Registros** | Datos crudos completos de la hoja DETALLE_HE |
                """)
                
                if st.button("⬇️ Generar y Descargar Reporte Excel", type="primary"):
                    with st.spinner("Generando reporte en formato Fortidex..."):
                        excel_data = generate_overtime_report_excel(
                            filtered_overtime_report, report_fecha_inicio, report_fecha_fin
                        )
                        
                        if excel_data:
                            fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"reporte_justificacion_sobretiempo_{fecha_str}.xlsx"
                            st.download_button(
                                label="💾 Descargar Reporte Oficial Fortidex",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.success("✅ Reporte generado en formato Justificación de Sobretiempo Fortidex. Haz clic para descargar.")
                        else:
                            st.error("❌ No se pudo generar el reporte. Verifica que haya datos disponibles.")
            else:
                st.warning("⚠️ No hay datos de horas extras para el período seleccionado.")
    else:
        st.info("Por favor, carga datos para comenzar.")

if __name__ == "__main__":
    main()
